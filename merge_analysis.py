# -*- coding: utf-8 -*-
"""
合并AI精排结果 -> 最终推荐 Excel

读取：
  analysis_inbox/pending_jobs.json     （job_matcher.py 生成的规则分 + 岗位信息）
  analysis_inbox/analysis_result.json  （AI写回的 match_score/reason/risk/verdict）
产出：
  景德镇_岗位精排推荐_<时间戳>.xlsx     （有AI结果时按 match_score 降序）
  景德镇_岗位规则推荐_<时间戳>.xlsx     （没有AI结果时按 rule_score 兜底排序）

最终排序分 = match_score(AI, 0-100) 为主；AI没覆盖到的岗位用规则分兜底。
如果还没写回 analysis_result.json，也可以先生成规则兜底推荐表。
"""

import datetime as dt
import json
import os
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INBOX = os.path.join(BASE_DIR, "analysis_inbox")

VERDICT_ORDER = {"强烈推荐": 0, "可投": 1, "观望": 2, "不建议": 3, "": 4}


def configure_console_encoding():
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass


def load_json(path, default):
    if not os.path.exists(path):
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"  ⚠️ 读取 {os.path.basename(path)} 失败：{e}")
        return default


def style_excel(path):
    wb = load_workbook(path)
    ws = wb.active
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    left_cols = set()
    link_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value in ("推荐理由", "风险提醒", "原始摘要", "岗位链接"):
            left_cols.add(idx)
        if cell.value == "岗位链接":
            link_col = idx
    link_font = Font(color="0563C1", underline="single")
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.border = border
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
            else:
                cell.alignment = left if cell.column in left_cols else center
                # 岗位链接列做成可直接点击的超链接(蓝色下划线)
                if cell.column == link_col and isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = cell.value
                    cell.font = link_font
    widths = {"类型": 10, "评级": 10, "精排分": 9, "规则分": 9, "岗位名称": 26, "公司名称": 24,
              "平台": 10, "薪资范围": 13, "工作地点": 16, "经验要求": 11, "学历要求": 10,
              "活跃度": 14, "推荐理由": 40, "风险提醒": 34, "岗位链接": 42}
    for idx, cell in enumerate(ws[1], start=1):
        col = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[col].width = widths.get(cell.value, 14)
    ws.freeze_panes = "A2"
    wb.save(path)


def main():
    configure_console_encoding()
    pending = load_json(os.path.join(INBOX, "pending_jobs.json"), {})
    result = load_json(os.path.join(INBOX, "analysis_result.json"), None)

    jobs = pending.get("岗位", []) if isinstance(pending, dict) else []
    if not jobs:
        print("  ✗ 没有 pending_jobs.json，先运行 job_matcher.py。")
        return
    if result is None:
        print("  ⚠️ 没找到 analysis_inbox/analysis_result.json。")
        print("    将先按规则分生成兜底推荐表；后续写回AI结果后可再次合并精排表。")
        result = []

    # 支持AI把结果包在 {"results":[...]} 或直接是数组
    if isinstance(result, dict):
        result = result.get("results") or result.get("岗位") or []
    by_id = {str(r.get("job_id")): r for r in result if isinstance(r, dict)}

    rows = []
    for j in jobs:
        jid = str(j.get("job_id"))
        a = by_id.get(jid, {})
        match_score = a.get("match_score")
        rule_score = j.get("rule_score", 0) or 0
        fallback_verdict = "强烈推荐" if rule_score >= 75 else "可投" if rule_score >= 55 else "观望"
        fallback_reason = "规则兜底排序，建议再用AI精排补充语义判断。"
        fallback_risk = "未经过AI精排，请人工确认工作时长、驻场/出差和薪资口径。"
        if j.get("hits"):
            fallback_reason = "规则命中：" + str(j.get("hits"))[:120]
        rows.append({
            "类型": j.get("类型", ""),
            "评级": a.get("verdict", fallback_verdict),
            "精排分": match_score if match_score is not None else "",
            "规则分": rule_score,
            "岗位名称": j.get("岗位名称", ""),
            "公司名称": j.get("公司名称", ""),
            "平台": j.get("平台", ""),
            "薪资范围": j.get("薪资范围", ""),
            "工作地点": j.get("工作地点", ""),
            "经验要求": j.get("经验要求", ""),
            "学历要求": j.get("学历要求", ""),
            "活跃度": j.get("活跃度", ""),
            "推荐理由": a.get("reason", fallback_reason),
            "风险提醒": a.get("risk", fallback_risk),
            "岗位链接": j.get("岗位链接", ""),
            "_sort_match": match_score if isinstance(match_score, (int, float)) else -1,
            "_sort_verdict": VERDICT_ORDER.get(a.get("verdict", fallback_verdict), 4),
            "_sort_rule": rule_score,
        })

    covered = sum(1 for r in rows if r["_sort_match"] >= 0)
    rows.sort(key=lambda r: (r["_sort_verdict"], -r["_sort_match"], -r["_sort_rule"]))
    df = pd.DataFrame(rows).drop(columns=["_sort_match", "_sort_verdict", "_sort_rule"])

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = "景德镇_岗位精排推荐" if covered else "景德镇_岗位规则推荐"
    out = os.path.join(BASE_DIR, f"{prefix}_{stamp}.xlsx")
    df.to_excel(out, index=False)
    style_excel(out)

    print(f"  ✅ 已合并 {len(rows)} 条岗位，其中 {covered} 条有AI精排。")
    print(f"  最终推荐表：{out}")
    if not covered:
        print("  提示：这是规则兜底结果。写回 analysis_result.json 后再运行，可生成AI精排表。")
    elif covered < len(rows):
        print(f"  注意：{len(rows) - covered} 条AI未给分（按规则分兜底排在后面）。")


if __name__ == "__main__":
    main()
