# -*- coding: utf-8 -*-
"""
====================================================
  BOSS直聘 回信监控 - DrissionPage 版 v1.0
====================================================

做什么：
  打开 BOSS 消息页，读出所有对话，找出"HR 已经回信"的，
  区分【对你投的岗位的回信】和【其它消息(常是广告/骚扰)】，
  并标出"上次检查之后的新回信"。结果存到 回信记录.xlsx。

判断逻辑：
  - 最后一条消息以 [送达]/[已读] 等方括号开头 = 我们发的，还没回 → 等待中
  - 最后一条不带方括号 = HR 发的 → 已回信

用法（连端口 9222 的调试 Chrome，需登录 BOSS）：
  python reply_monitor.py
"""

import datetime as dt
import json
import os
import re
import sys
import time
import traceback

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)

try:
    from notify import notify
except Exception:
    def notify(title, msg):
        pass
RECORD_PATH = os.path.join(BASE_DIR, "投递记录.json")
SNAPSHOT_PATH = os.path.join(BASE_DIR, "回信快照.json")
OUT_XLSX = os.path.join(BASE_DIR, "回信记录.xlsx")


def log(msg):
    print(msg, flush=True)


def connect_chrome(port=9222):
    from DrissionPage import ChromiumPage, ChromiumOptions
    co = ChromiumOptions()
    co.set_local_port(port)
    return ChromiumPage(co)


def load_applied_companies():
    if not os.path.exists(RECORD_PATH):
        return set()
    try:
        with open(RECORD_PATH, "r", encoding="utf-8") as f:
            recs = json.load(f)
        return {str(r.get("公司名称", "")).strip() for r in recs if r.get("公司名称")}
    except Exception:
        return set()


# ---- 公司名匹配 ----
# 投递记录里的公司名常被 BOSS 列表页截断（如"景德镇市陶阳里景..."），
# 直接拿整串去聊天列表里找永远找不到 → 真回信被误标成广告。
# 处理：去掉末尾省略号，再拆出"去行政区前缀 / 去公司后缀"的核心名，任一片段命中即算匹配。

_REGION_PREFIX = re.compile(
    r"^(?:江西|江苏|浙江|广东|山东|湖南|湖北|安徽|福建|云南|四川|贵州|河南|河北|"
    r"陕西|山西|广西|海南|辽宁|吉林|黑龙江|甘肃|青海|内蒙古|宁夏|新疆|西藏|"
    r"北京|上海|天津|重庆|[一-龥]{1,3}(?:省|市|区|县))"
)
_COMPANY_SUFFIX = re.compile(r"(?:股份)?(?:有限)?(?:责任)?公司$|工作室$|事务所$")


def _company_keys(name):
    """公司名 → 一组可匹配的片段。片段至少 3 个字，避免误伤（完整名放宽到 2 字）。"""
    base = re.sub(r"[.…·\s]+$", "", str(name or "").strip())
    keys = set()
    if len(base) >= 2:
        keys.add(base)
    stripped = base
    for _ in range(3):  # 前缀可叠多层：江西 / 景德镇市 / 珠山区
        m = _REGION_PREFIX.match(stripped)
        if not m or len(stripped) - (m.end() - m.start()) < 3:
            break
        stripped = stripped[m.end():]
        keys.add(stripped)
    for k in list(keys):
        core = _COMPANY_SUFFIX.sub("", k)
        if len(core) >= 3:
            keys.add(core)
    return keys


def company_match(applied_names, who):
    """聊天列表里的对方描述 who 是否命中任何一家投过的公司。"""
    who = str(who or "")
    if not who:
        return False
    for comp in applied_names:
        for key in _company_keys(comp):
            if key in who:
                return True
    return False


def load_snapshot():
    if os.path.exists(SNAPSHOT_PATH):
        try:
            with open(SNAPSHOT_PATH, "r", encoding="utf-8") as f:
                return set(json.load(f))
        except Exception:
            return set()
    return set()


def save_snapshot(keys):
    with open(SNAPSHOT_PATH, "w", encoding="utf-8") as f:
        json.dump(sorted(keys), f, ensure_ascii=False, indent=2)


def read_conversations(tab):
    """读消息列表，返回 [{time, who, last, replied}]。"""
    ul = tab.ele(".user-list", timeout=8)
    if not ul:
        return []
    convos = []
    for li in ul.eles("tag:li"):
        try:
            lines = [x.strip() for x in (li.text or "").split("\n") if x.strip()]
        except Exception:
            continue  # 个别条目读取时页面刷新导致元素失效，跳过不影响整体
        if not lines:
            continue
        # 有未读消息时第一行是未读数角标（如"1"），会把后面的列全挤错位，先摘掉
        unread = 0
        if re.fullmatch(r"\d{1,3}", lines[0]):
            unread = int(lines.pop(0))
        if not lines:
            continue
        ctime = lines[0] if lines else ""
        who = lines[1] if len(lines) > 1 else ""
        last = lines[2] if len(lines) > 2 else ""
        # 有未读=对方发了新消息；否则看最后一条是否以 [送达]/[已读] 等方括号开头（是=我们发的）
        replied = unread > 0 or (bool(last) and not last.startswith("["))
        convos.append({"time": ctime, "who": who, "last": last, "replied": replied})
    return convos


def save_xlsx(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "回信监控"
    headers = ["时间", "对方(HR·公司·职位)", "最后一条消息", "类别", "是否新回信"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([r["time"], r["who"], r["last"], r["category"], "新" if r["is_new"] else ""])
    widths = [10, 34, 50, 16, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    # 高亮新回信
    green = PatternFill("solid", fgColor="E8FFEA")
    for ri, r in enumerate(rows, 2):
        if r["is_new"]:
            for ci in range(1, 6):
                ws.cell(ri, ci).fill = green
    wb.save(OUT_XLSX)


def run_check(page, header=True):
    """用已有的 page 跑一次回信检查（供 auto_apply 投递中调用，不必重连）。"""
    if header:
        log("=" * 55)
        log("  BOSS直聘 回信监控")
        log("=" * 55)
    tab = None
    for attempt in (1, 2):
        try:
            tab = page.new_tab("https://www.zhipin.com/web/geek/chat")
            break
        except Exception as e:
            if attempt == 2:
                raise
            log(f"  (打开消息页失败，3秒后重试一次：{e})")
            time.sleep(3)
    time.sleep(5)
    if "login" in (tab.url or ""):
        log("  ⚠️ 看起来没登录，请在 Chrome 里登录 BOSS 后重试。")
        try:
            tab.close()
        except Exception:
            pass
        return

    convos = read_conversations(tab)
    try:
        tab.close()
    except Exception:
        pass

    if not convos:
        log("  没读到对话（可能页面还没加载好，或暂无消息）。")
        return

    applied = load_applied_companies()
    snapshot = load_snapshot()

    rows = []
    new_replies = 0
    yours_replies = 0
    new_yours = 0
    for c in convos:
        if not c["replied"]:
            continue
        is_yours = company_match(applied, c["who"])
        category = "✅ 你投的岗位回信" if is_yours else "对方主动找你(可能是机会/也可能是广告)"
        key = c["who"] + "|" + c["last"]
        is_new = key not in snapshot
        rows.append({**c, "category": category, "is_new": is_new})
        if is_new:
            new_replies += 1
            if is_yours:
                new_yours += 1
        if is_yours:
            yours_replies += 1

    # 更新快照（记录所有当前回信，下次比对）
    save_snapshot({r["who"] + "|" + r["last"] for r in rows})

    log(f"  共 {len(convos)} 个对话；其中已回信 {len(rows)} 个；"
        f"对你投的岗位回信 {yours_replies} 个；本次新增回信 {new_replies} 个")
    log("-" * 55)
    if not rows:
        log("  暂时还没有 HR 回信，过会儿再看看。")
    else:
        # 先显示新回信、你投的优先
        rows.sort(key=lambda r: (not r["is_new"], "你投的" not in r["category"]))
        for r in rows:
            flag = "🆕 " if r["is_new"] else "   "
            log(f"  {flag}{r['category']} | {r['who']}")
            log(f"      「{r['last'][:60]}」 ({r['time']})")
    log("-" * 55)
    try:
        save_xlsx(rows)
        log(f"  📁 已保存：{os.path.basename(OUT_XLSX)}")
    except Exception as e:
        log(f"  (保存 Excel 失败：{e})")
    if new_replies:
        log(f"  🔔 有 {new_replies} 条新回信！记得去 BOSS 回复。")
        if new_yours:
            notify("求职助手：你投的岗位有回信了！",
                   f"{new_yours} 条来自你投过的公司，快去 BOSS 回复，别晾着 HR。")
        else:
            notify("求职助手：有新消息",
                   f"{new_replies} 条新消息（对方主动找你，可能是机会也可能是广告），有空瞄一眼。")
    log("=" * 55)


def main():
    log("=" * 55)
    log("  BOSS直聘 回信监控")
    log("=" * 55)
    log("  连接调试 Chrome（端口 9222）……")
    try:
        page = connect_chrome()
    except Exception as e:
        log(f"  ❌ 连不上 Chrome：{e}")
        log("     请先点【②打开网站并登录】启动调试 Chrome 并登录 BOSS。")
        return
    try:
        run_check(page, header=False)
    except Exception:
        # 报错原因完整打出来，让总控日志里查得到，别再无声失败
        log("  ❌ 回信检查出错，原因如下：")
        log(traceback.format_exc())
        sys.exit(1)


if __name__ == "__main__":
    main()
