# -*- coding: utf-8 -*-
r"""
全平台一键抓取（BOSS + 51job + 智联，可扩展）
依次抓取各平台 → 抽取公共列、标注来源平台、跨平台去重 → 合并成一张
『全平台_岗位采集_<时间>.xlsx』，供现有 job_matcher → AI精排 → 出表/网页 统一使用。

注意：
  - 每个平台都需在调试 Chrome(9222) 里**各自登录**；本脚本只负责抓取与合并。
  - 自动投递目前仍只支持 BOSS；其它平台在推荐里给链接，手动投。
用法：
  python 全平台抓取.py --city 景德镇 --keywords "电商运营助理,数据助理" --platforms boss,51job,zhaopin
"""
import argparse
import datetime as dt
import glob
import os
import subprocess
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

from openpyxl import load_workbook, Workbook

BASE = os.path.dirname(os.path.abspath(__file__))
PY = sys.executable

# 平台 -> (中文名, 采集文件名前缀)
PLAT = {
    "51job": "前程无忧",
    "zhaopin": "智联招聘",
    "liepin": "猎聘",
    "lagou": "拉勾",
}
BOSS_OUT = "全平台BOSS采集.xlsx"
# 各平台统一抽取的公共列
COMMON = ["平台", "岗位名称", "薪资范围", "公司名称", "工作地点",
          "经验要求", "学历要求", "岗位链接", "搜索关键词"]


def log(m):
    print(m, flush=True)


def run(args, title):
    log("-" * 55)
    log(f"▶ {title}")
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    p = subprocess.Popen([PY] + args, cwd=BASE, stdout=subprocess.PIPE,
                         stderr=subprocess.STDOUT, text=True, encoding="utf-8",
                         errors="replace", env=env)
    for line in p.stdout:
        line = line.rstrip()
        if line:
            print("   " + line, flush=True)
    return p.wait()


def latest(pattern):
    fs = [f for f in glob.glob(os.path.join(BASE, pattern))
          if not os.path.basename(f).startswith("~$")]
    return max(fs, key=os.path.getmtime) if fs else None


def read_rows(path, platform_label=None):
    """读一个采集文件，抽公共列，返回 dict 列表。"""
    if not path or not os.path.exists(path):
        return []
    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        d = dict(zip(headers, r))
        rec = {c: ("" if d.get(c) is None else d.get(c, "")) for c in COMMON}
        if platform_label and not str(rec.get("平台", "")).strip():
            rec["平台"] = platform_label
        if str(rec.get("岗位名称", "")).strip():
            out.append(rec)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--city", default="景德镇")
    ap.add_argument("--keywords", default="电商运营助理,数据助理")
    ap.add_argument("--platforms", default="boss,51job,zhaopin")
    ap.add_argument("--pages", type=int, default=2)
    args = ap.parse_args()
    plats = [p.strip() for p in args.platforms.split(",") if p.strip()]

    log("=" * 55)
    log(f"  全平台抓取：{plats}  城市={args.city}")
    log("=" * 55)

    collected = {}  # platform -> file path

    if "boss" in plats:
        run(["boss_scraper.py", "--city", args.city, "--max-pages", str(args.pages),
             "--max-details", "0", "--skip-reputation", "--yes",
             "--keywords", args.keywords, "--no-strict-match", "--fast",
             "--output", BOSS_OUT], "抓取 BOSS直聘")
        collected["BOSS直聘"] = latest("全平台BOSS采集*.xlsx")

    for key in plats:
        if key == "boss":
            continue
        name = PLAT.get(key, key)
        run(["multi_platform_scraper.py", "--platform", key, "--city", args.city,
             "--keywords", args.keywords, "--pages", str(args.pages), "--yes"],
            f"抓取 {name}")
        collected[name] = latest(f"{name}_*岗位采集*.xlsx")

    # 合并 + 去重
    log("-" * 55)
    log("  合并各平台数据……")
    seen = set()
    merged = []
    per = {}
    for label, path in collected.items():
        rows = read_rows(path, platform_label=label)
        per[label] = len(rows)
        for rec in rows:
            link = str(rec.get("岗位链接", "")).strip()
            key = link or (str(rec.get("公司名称", "")) + "|" + str(rec.get("岗位名称", "")))
            if key in seen:
                continue
            seen.add(key)
            merged.append(rec)

    log("  各平台条数：" + " ".join(f"{k}={v}" for k, v in per.items()))
    if not merged:
        log("  ❌ 没合并到任何岗位（可能各平台都没登录/没结果）。")
        return

    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    out = os.path.join(BASE, f"{args.city}_全平台_岗位采集_{ts}.xlsx")
    wb = Workbook()
    sh = wb.active
    sh.title = "全平台岗位"
    sh.append(COMMON)
    for rec in merged:
        sh.append([rec.get(c, "") for c in COMMON])
    wb.save(out)
    log("=" * 55)
    log(f"  ✅ 合并完成：共 {len(merged)} 条（已跨平台去重）")
    log(f"  📁 {os.path.basename(out)}")
    log("  下一步：点『④ AI 筛选并出推荐表』，就会对这张全平台表统一精排。")
    log("=" * 55)


if __name__ == "__main__":
    main()
