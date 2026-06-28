# -*- coding: utf-8 -*-
"""把最新推荐表 Excel 生成一个好看的、可筛选/排序的单文件网页(双击即可在浏览器查看)。"""
import json, glob, os, sys
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))


def latest_table():
    files = glob.glob(os.path.join(BASE, "景德镇_岗位*推荐_*.xlsx"))
    return max(files, key=os.path.getmtime) if files else None


TEMPLATE = r"""<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>__TITLE__</title>
<style>
* { box-sizing:border-box; margin:0; padding:0; }
body { font-family:"Microsoft YaHei","PingFang SC",sans-serif; background:#f4f5f7; color:#1f2329; padding:20px; }
.wrap { max-width:1180px; margin:0 auto; }
h1 { font-size:22px; color:#1f3864; }
.sub { color:#86909c; font-size:13px; margin-top:5px; line-height:1.6; }
.stats { display:flex; gap:10px; margin-top:12px; flex-wrap:wrap; }
.chip { background:#fff; border:1px solid #e5e6eb; border-radius:20px; padding:5px 14px; font-size:13px; }
.chip b { color:#1f3864; font-size:15px; }
.toolbar { background:#fff; border-radius:10px; padding:12px 16px; margin:16px 0; display:flex; gap:16px; flex-wrap:wrap; align-items:center; box-shadow:0 1px 3px rgba(0,0,0,.05); }
.group { display:flex; gap:6px; align-items:center; }
.group label { font-size:13px; color:#4e5969; margin-right:2px; }
.btn { border:1px solid #e5e6eb; background:#fff; border-radius:6px; padding:5px 12px; font-size:13px; cursor:pointer; color:#4e5969; }
.btn.active { background:#1f3864; color:#fff; border-color:#1f3864; }
input,select { border:1px solid #e5e6eb; border-radius:6px; padding:6px 10px; font-size:13px; outline:none; }
input:focus,select:focus { border-color:#1f3864; }
.grid { display:grid; grid-template-columns:repeat(auto-fill,minmax(355px,1fr)); gap:14px; }
.card { background:#fff; border-radius:10px; padding:16px; box-shadow:0 1px 3px rgba(0,0,0,.06); transition:.15s; border:1px solid #f0f0f0; display:flex; flex-direction:column; }
.card:hover { box-shadow:0 6px 18px rgba(0,0,0,.12); transform:translateY(-2px); }
.top { display:flex; justify-content:space-between; align-items:flex-start; gap:8px; }
.jt { font-size:16px; font-weight:700; }
.sal { font-size:16px; font-weight:700; color:#f53f3f; white-space:nowrap; }
.badges { display:flex; gap:6px; flex-wrap:wrap; margin:10px 0; }
.bdg { font-size:12px; padding:2px 9px; border-radius:4px; font-weight:500; }
.b-grow { background:#f3eafd; color:#7b3ff2; }
.b-easy { background:#fff3e8; color:#f77234; }
.b-strong { background:#e8ffea; color:#00b42a; }
.b-ok { background:#e8f3ff; color:#165dff; }
.b-watch { background:#f2f3f5; color:#86909c; }
.b-no { background:#ffece8; color:#f53f3f; }
.b-loc { background:#f2f3f5; color:#4e5969; }
.b-active { background:#e8fffb; color:#0aa6a2; }
.b-plat { background:#1f3864; color:#fff; }
.company { font-size:13px; color:#4e5969; margin-bottom:8px; }
.reason { font-size:13px; color:#1d8c4d; background:#f0faf4; padding:7px 10px; border-radius:6px; margin-bottom:6px; line-height:1.5; }
.risk { font-size:13px; color:#cb6a14; background:#fff7e8; padding:7px 10px; border-radius:6px; margin-bottom:10px; line-height:1.5; }
.meta { font-size:12px; color:#a9aeb8; margin-bottom:12px; }
.go { margin-top:auto; text-align:center; background:#1f3864; color:#fff; text-decoration:none; padding:9px; border-radius:6px; font-size:14px; font-weight:600; }
.go:hover { background:#16305a; }
.empty { text-align:center; color:#86909c; padding:60px; grid-column:1/-1; }
.foot { text-align:center; color:#86909c; font-size:13px; margin:26px 0 10px; line-height:1.9; }
.foot a { color:#1f3864; text-decoration:none; font-weight:600; } .foot a:hover { text-decoration:underline; }
</style>
</head>
<body>
<div class="wrap">
<h1>__TITLE__</h1>
<div class="sub">按契合度智能排序 · 已过滤高薪诈骗 / 僵尸号 / 重体力 · 标注成长型与轻松型 · 离家近优先</div>
<div class="stats" id="stats"></div>
<div class="toolbar">
<div class="group"><label>类型</label>
<button class="btn active" data-type="">全部</button>
<button class="btn" data-type="成长型">成长型</button>
<button class="btn" data-type="轻松型">轻松型</button></div>
<div class="group"><label>评级</label>
<button class="btn active" data-verd="">全部</button>
<button class="btn" data-verd="强烈推荐">强烈推荐</button>
<button class="btn" data-verd="可投">可投</button></div>
<div class="group"><label>排序</label>
<select id="sort"><option value="score">按推荐分</option><option value="salary">按薪资</option></select></div>
<div class="group" style="flex:1"><input id="search" placeholder="搜岗位名 / 公司…" style="width:100%"></div>
</div>
<div class="grid" id="grid"></div>
<div class="foot">本工具由 B站 <b>@__AUTHOR__</b> 原创开发 · <a href="__BILI__" target="_blank" rel="noopener">▶ B站主页</a> · <a href="__GH__" target="_blank" rel="noopener">★ GitHub 开源</a></div>
</div>
<script>
const DATA = __DATA__;
let fType="", fVerd="", fSort="score", fSearch="";
function salaryNum(s){ const m=String(s||"").match(/(\d+(?:\.\d+)?)/); return m?parseFloat(m[1]):0; }
function scoreNum(d){ return parseFloat(d["精排分"]||d["规则分"]||0)||0; }
function verdClass(v){ return v==="强烈推荐"?"b-strong":v==="可投"?"b-ok":v==="不建议"?"b-no":"b-watch"; }
function esc(s){ const d=document.createElement("div"); d.textContent=(s==null?"":String(s)); return d.innerHTML; }
function render(){
  let list = DATA.filter(d=>{
    if(fType && d["类型"]!==fType) return false;
    if(fVerd && d["评级"]!==fVerd) return false;
    if(fSearch){ const t=(String(d["岗位名称"])+String(d["公司名称"])).toLowerCase(); if(!t.includes(fSearch.toLowerCase())) return false; }
    return true;
  });
  list.sort((a,b)=> fSort==="salary" ? salaryNum(b["薪资范围"])-salaryNum(a["薪资范围"]) : scoreNum(b)-scoreNum(a));
  const grid=document.getElementById("grid");
  if(!list.length){ grid.innerHTML='<div class="empty">没有符合条件的岗位</div>'; return; }
  grid.innerHTML = list.map(d=>{
    const t=d["类型"], v=d["评级"], link=d["岗位链接"]||"";
    const reason=d["推荐理由"]?'<div class="reason">✓ '+esc(d["推荐理由"])+'</div>':"";
    const risk=d["风险提醒"]?'<div class="risk">⚠ '+esc(d["风险提醒"])+'</div>':"";
    const tB = t==="成长型"?'<span class="bdg b-grow">成长型</span>':t==="轻松型"?'<span class="bdg b-easy">轻松型</span>':"";
    const act = d["活跃度"]?'<span class="bdg b-active">'+esc(d["活跃度"])+'</span>':"";
    const plat = d["平台"]?'<span class="bdg b-plat">'+esc(d["平台"])+'</span>':"";
    const isBoss = /zhipin\.com/.test(link);
    const btnText = isBoss ? "查看岗位 / 立即沟通 →" : "去该平台投递 →";
    return '<div class="card"><div class="top"><div class="jt">'+esc(d["岗位名称"])+'</div><div class="sal">'+esc(d["薪资范围"]||"面议")+'</div></div>'+
      '<div class="badges">'+plat+tB+'<span class="bdg '+verdClass(v)+'">'+esc(v)+'</span><span class="bdg b-loc">'+esc(d["工作地点"]||"")+'</span>'+act+'</div>'+
      '<div class="company">'+esc(d["公司名称"]||"")+' · '+esc(d["经验要求"]||"经验不限")+' · '+esc(d["学历要求"]||"")+'</div>'+
      reason+risk+'<div class="meta">推荐分 '+esc(d["精排分"]||d["规则分"]||"")+'</div>'+
      (link?'<a class="go" href="'+esc(link)+'" target="_blank" rel="noopener">'+btnText+'</a>':'')+'</div>';
  }).join("");
}
function renderStats(){
  const g=DATA.filter(d=>d["类型"]==="成长型").length, e=DATA.filter(d=>d["类型"]==="轻松型").length;
  document.getElementById("stats").innerHTML='<span class="chip">共 <b>'+DATA.length+'</b> 个岗位</span><span class="chip">成长型 <b>'+g+'</b></span><span class="chip">轻松型 <b>'+e+'</b></span>';
}
document.querySelectorAll("[data-type]").forEach(b=>b.onclick=()=>{document.querySelectorAll("[data-type]").forEach(x=>x.classList.remove("active"));b.classList.add("active");fType=b.dataset.type;render();});
document.querySelectorAll("[data-verd]").forEach(b=>b.onclick=()=>{document.querySelectorAll("[data-verd]").forEach(x=>x.classList.remove("active"));b.classList.add("active");fVerd=b.dataset.verd;render();});
document.getElementById("sort").onchange=e=>{fSort=e.target.value;render();};
document.getElementById("search").oninput=e=>{fSearch=e.target.value;render();};
renderStats(); render();
</script>
</body>
</html>"""


def main():
    f = latest_table()
    if not f:
        print("没有推荐表,先用 ④ 出一份推荐表。")
        return
    wb = load_workbook(f)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append({h: ("" if v is None else v) for h, v in zip(headers, row)})
    # 标题动态生成、不含个人姓名（保护隐私；城市跟着 user_profile 走）
    city = "本地"
    prof_path = os.path.join(BASE, "user_profile.json")
    if os.path.exists(prof_path):
        try:
            with open(prof_path, "r", encoding="utf-8") as pf:
                city = json.load(pf).get("city") or city
        except Exception:
            pass
    title = f"{city}岗位推荐"
    html = (TEMPLATE.replace("__DATA__", json.dumps(data, ensure_ascii=False))
            .replace("__TITLE__", title)
            .replace("__AUTHOR__", "1285452862")
            .replace("__BILI__", "https://space.bilibili.com/352671558")
            .replace("__GH__", "https://github.com/hdrtfhyrs/local-job-assistant"))
    out = os.path.join(os.path.expanduser("~"), "Desktop", "求职推荐_网页版.html")
    with open(out, "w", encoding="utf-8") as fp:
        fp.write(html)
    print("SAVED:", out)
    print("岗位数:", len(data))
    if data:
        print("列:", list(data[0].keys()))
        print("样例:", json.dumps(data[0], ensure_ascii=False)[:200])


if __name__ == "__main__":
    main()
