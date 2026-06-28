# -*- coding: utf-8 -*-
"""
====================================================
  BOSS直聘 自动投递 (打招呼) - DrissionPage 版 v1.0
====================================================

做什么：
  读取最新的【岗位精排推荐】表，按推荐分从高到低，
  对符合条件、且没投过的岗位，自动打开岗位页 → 点"立即沟通" → 发招呼语。

保命措施（防封号 / 防误投）：
  - 每日上限（默认 60 条/天，超了自动停）
  - 每条之间随机间隔（默认 40~180 秒，模拟真人）
  - 去重：投过的岗位不再投（投递记录.json）
  - 避坑：评级"不建议"、风险提醒命中黑名单关键词的，自动跳过
  - 预演模式 --dry-run：只定位"立即沟通"按钮、不真的发，用来先校准点位

用法（都连端口 9222 的调试 Chrome，需先登录 BOSS）：
  预演（强烈建议先跑这个）：  python auto_apply.py --dry-run
  真投：                      python auto_apply.py
  限量真投：                  python auto_apply.py --max 30
"""

import argparse
import datetime as dt
import json
import os
import random
import re
import sys
import time

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RECORD_PATH = os.path.join(BASE_DIR, "投递记录.json")
CONFIG_PATH = os.path.join(BASE_DIR, "投递配置.json")

# ---------------- 默认配置（投递配置.json 存在时会覆盖） ----------------
DEFAULTS = {
    "greeting": (
        "您好，我在BOSS直聘看到贵公司的这个岗位，很感兴趣。我本科软件工程毕业，"
        "学习能力强、踏实肯干，希望在本地长期稳定发展。方便的话期待和您进一步沟通，谢谢！"
    ),
    "daily_cap": 200,           # 每天最多投多少（含今天已投的）；设高=投到BOSS拦/岗位投完为止
    "min_grade": "观望",        # 投递门槛：达到这个评级及以上才投（广撒网就放到"观望"）
    "delay_min": 25,            # 每条之间最少间隔(秒)
    "delay_max": 90,            # 每条之间最多间隔(秒)
    "monitor_every": 3,         # 每投几个就跑一次回信监控（0=关闭）
    "trap_keywords": ["黑名单", "诈骗", "传销", "僵尸", "虚假", "已关闭", "押金", "刷单", "兼职日结"],
    # 岗位名称命中这些=要专业证书/手艺、与软件工程背景不对口，直接跳过(不浪费额度)
    "skip_title_keywords": [
        "财务", "会计", "出纳", "审计", "税务", "金融", "证券", "保险",
        "电工", "焊工", "钳工", "车工", "铣工", "数控", "机修", "维修电工",
        "电气", "电力", "电站", "变电", "高压", "钣金", "喷漆", "油漆",
        "木工", "瓦工", "钢筋", "架子工", "水电", "管道", "焊接", "锅炉",
        "护士", "护理", "医生", "医师", "药剂", "药师", "检验", "麻醉",
        "律师", "法务", "教师", "老师", "幼师", "厨师", "面点", "烘焙", "厨房",
        "造价", "预算", "施工员", "测量", "测绘", "翻译", "兽医", "美容", "美发",
        "理疗", "推拿", "按摩", "司机", "驾驶", "保安", "保洁", "月嫂", "育儿",
        "设计师", "主播", "直播", "教练", "中介", "置业顾问", "房产经纪",
    ],
}

# 评级从高到低的等级（数字越大越好），用于门槛比较
GRADE_RANK = {"不建议": 0, "观望": 1, "可投": 2, "强烈推荐": 3}


def log(msg):
    print(msg, flush=True)


PROFILE_PATH = os.path.join(BASE_DIR, "user_profile.json")


def load_config():
    cfg = dict(DEFAULTS)
    # 1) 个性化画像 user_profile.json：用它生成的"不对口黑名单/招呼语"覆盖默认
    #    （这样会计用就投财务、软件用就跳财务——按各人背景来，而不是写死）
    if os.path.exists(PROFILE_PATH):
        try:
            with open(PROFILE_PATH, "r", encoding="utf-8") as f:
                prof = json.load(f)
            gen = prof.get("generated") or {}
            off = gen.get("off_fit_keywords")
            if off:
                cfg["skip_title_keywords"] = off
            if gen.get("greeting"):
                cfg["greeting"] = gen["greeting"]
            log(f"  （已按你的个性化画像 user_profile.json 配置投递规则）")
        except Exception as e:
            log(f"  (user_profile.json 读取失败，用默认: {e})")
    # 2) 投递配置.json：细项覆盖（上限/间隔/门槛等），优先级最高
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                user = json.load(f)
            cfg.update({k: v for k, v in user.items() if v is not None})
        except Exception as e:
            log(f"  (投递配置.json 读取失败，用默认配置: {e})")
    return cfg


def latest_table():
    """找最新的精排推荐表；没有就退而求其次找规则推荐表。"""
    import glob
    for pat in ("景德镇_岗位精排推荐_*.xlsx", "*岗位精排推荐*.xlsx",
                "*岗位规则推荐*.xlsx", "*自定义岗位搜集*.xlsx"):
        files = [f for f in glob.glob(os.path.join(BASE_DIR, pat))
                 if not os.path.basename(f).startswith("~$")]
        if files:
            return max(files, key=os.path.getmtime)
    return None


def job_id_of(link):
    """从岗位链接里抽出唯一 id，用于去重。"""
    if not link:
        return ""
    m = re.search(r"job_detail/([0-9a-zA-Z~_-]+)", str(link))
    return m.group(1) if m else str(link).strip()


def load_records():
    if os.path.exists(RECORD_PATH):
        try:
            with open(RECORD_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []


def save_records(records):
    with open(RECORD_PATH, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)


def today_str():
    return dt.datetime.now().strftime("%Y-%m-%d")


def read_jobs(table_path):
    """读表 → 字典列表。"""
    wb = load_workbook(table_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    jobs = []
    for r in rows[1:]:
        d = {h: ("" if v is None else v) for h, v in zip(headers, r)}
        if any(str(v).strip() for v in d.values()):
            jobs.append(d)
    return jobs


# 风险提醒里出现这些"劝告语气"时，命中的坑词多半只是提醒小心、不是确认是坑
ADVISORY_MARKERS = ("避免", "谨防", "以防", "是否", "确认", "关注", "留意", "防止", "建议", "注意", "核实", "需关注")


def is_trap(job, trap_keywords):
    """是不是该跳过的坑。返回(是否跳过, 原因)。"""
    grade = str(job.get("评级", "")).strip()
    if grade == "不建议":
        return True, "评级=不建议"
    risk = str(job.get("风险提醒", ""))
    # 若整句是劝告语气(如"建议关注HR,避免投到僵尸号")，不把坑词当真——交给评级和AI关卡判断
    if any(m in risk for m in ADVISORY_MARKERS):
        return False, ""
    for kw in trap_keywords:
        if kw and kw in risk:
            return True, f"风险提醒命中'{kw}'"
    return False, ""


def grade_ok(job, min_grade):
    grade = str(job.get("评级", "")).strip()
    need = GRADE_RANK.get(min_grade, 1)
    have = GRADE_RANK.get(grade, 1)  # 没评级的当"观望"对待
    return have >= need


def score_of(job):
    for k in ("精排分", "规则分"):
        try:
            return float(job.get(k) or 0)
        except (ValueError, TypeError):
            continue
    return 0.0


def load_profile_brief():
    """读 user_profile.json 拼一句求职者背景，给 AI 对口关卡用。没有则空串。"""
    p = os.path.join(BASE_DIR, "user_profile.json")
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8") as f:
                prof = json.load(f)
            return (f"城市{prof.get('city','')}，{prof.get('education','')}学历；"
                    f"背景：{prof.get('background','')}；想做的方向：{prof.get('targets','')}；"
                    f"不接受：{prof.get('avoid','')}")
        except Exception:
            pass
    return ""


def ai_fit_check(profile_brief, job, model, jd=""):
    """问大模型：结合公司行业，这岗位是否需要求职者不具备的专业技能/资质。
    返回 (fit:bool, reason)。出错一律按 fit=True（宁可投，不误杀）。"""
    try:
        from local_ai_matcher import call_ollama
        system = (
            "你判断给定岗位对这位求职者是否专业对口，只输出 JSON。判断准则：\n"
            "【算对口】信息技术/软件/系统类岗位——软件、系统、网络、IT、实施(含ERP/HIS等系统实施)、"
            "IT技术支持、测试、IT运维、数据，以及电商/运营/新媒体/客服/行政等通用岗位；"
            "这些即使在医疗/金融/文旅等行业也算对口。\n"
            "【算不对口】需要求职者画像中不具备的专业资质或现场专业技能，例如："
            "①环保/电力/化工/机械/建筑等工业企业里、名称【不含】'IT/信息/软件/系统/网络'字样的"
            "『运维/技术支持/工程师』(这类多为现场设备或工艺岗)；②电工/焊工/数控等技工；"
            "③会计(需会计证)、设计(需设计功底)、医护、施工/造价/测绘等。\n"
            "【特别注意】仔细看岗位描述里的硬性要求：若要求会某项求职者不具备的专业技能"
            "(如会CAD/制图、机械制图、特定编程语言深度经验、持某证书、某专业背景)，就算名称对口也判不对口。\n"
            "结合公司行业、岗位名称和岗位描述综合判断。"
            'JSON格式：{"fit": true或false, "reason": "20字内理由"}'
        )
        user = (
            f"【求职者】{profile_brief}\n"
            f"【岗位】{job.get('岗位名称','')} @ {job.get('公司名称','')}；"
            f"学历要求{job.get('学历要求','')}、经验{job.get('经验要求','')}；"
            f"推荐理由：{job.get('推荐理由','')}；已知风险：{job.get('风险提醒','')}\n"
            f"【岗位描述】{jd[:500] if jd else '(无)'}"
        )
        raw = call_ollama(model, system, user, timeout=60)
        data = raw if isinstance(raw, dict) else json.loads(raw)
        return bool(data.get("fit", True)), str(data.get("reason", ""))[:40]
    except Exception:
        return True, "(AI关卡异常,放行)"


def ai_guard_filter(cands, model):
    """投递前用 AI 逐个核查专业是否对口，剔除不对口的。返回 (保留, 跳过列表)。"""
    brief = load_profile_brief()
    if not brief:
        return cands, []  # 没填画像就不做这道关卡
    try:
        from local_ai_matcher import check_ollama
        if not check_ollama(model):
            log("  （本地大模型不可用，跳过 AI 对口核查，仅用名称黑名单）")
            return cands, []
    except Exception:
        return cands, []
    kept, dropped = [], []
    log(f"  投递前 AI 专业对口核查（共 {len(cands)} 个，用 {model}）……")
    for job, jid, link in cands:
        fit, reason = ai_fit_check(brief, job, model)
        if fit:
            kept.append((job, jid, link))
        else:
            dropped.append((job, reason))
            log(f"    ✗ 跳过(专业不对口) {str(job.get('岗位名称',''))[:16]}@{str(job.get('公司名称',''))[:12]} — {reason}")
    log(f"  AI 关卡完成：保留 {len(kept)} 个，剔除 {len(dropped)} 个")
    return kept, dropped


def build_candidates(jobs, records, cfg):
    """筛选 + 排序，得到本次要投的候选。"""
    applied_ids = {r.get("job_id") for r in records}
    skip_titles = cfg.get("skip_title_keywords", [])
    cands = []
    skipped = {"已投过": 0, "无链接": 0, "门槛不够": 0, "踩坑": 0, "专业不对口": 0}
    for job in jobs:
        link = job.get("岗位链接") or job.get("链接") or ""
        jid = job_id_of(link)
        if not jid:
            skipped["无链接"] += 1
            continue
        if jid in applied_ids:
            skipped["已投过"] += 1
            continue
        title = str(job.get("岗位名称", ""))
        if any(kw and kw in title for kw in skip_titles):
            skipped["专业不对口"] += 1
            continue
        trap, why = is_trap(job, cfg["trap_keywords"])
        if trap:
            skipped["踩坑"] += 1
            continue
        if not grade_ok(job, cfg["min_grade"]):
            skipped["门槛不够"] += 1
            continue
        cands.append((job, jid, link))
    # 推荐分从高到低，好的先投
    cands.sort(key=lambda x: score_of(x[0]), reverse=True)
    return cands, skipped


# ---------------- 浏览器自动化 ----------------

def connect_chrome(port):
    from DrissionPage import ChromiumPage, ChromiumOptions
    co = ChromiumOptions()
    co.set_local_port(port)
    page = ChromiumPage(co)
    return page


def _first_visible(eles):
    """从一堆元素里挑第一个真正可见(有位置和大小)的，避开隐藏的悬浮按钮。"""
    for e in eles or []:
        try:
            if e.states.is_displayed:
                w, h = e.rect.size
                if w > 0 and h > 0:
                    return e
        except Exception:
            continue
    return None


def read_jd(tab):
    """读岗位详情页的'职位描述'文本（让AI看真实硬要求，如会CAD/持证）。"""
    for sel in (".job-sec-text", ".job-detail-section", ".job-sec"):
        try:
            e = tab.ele(sel, timeout=1.5)
            if e:
                t = (e.text or "").strip()
                if len(t) > 8:
                    return t[:600]
        except Exception:
            continue
    return ""


def find_greet_button(tab):
    """在岗位详情页找'立即沟通'按钮。返回 (按钮, 状态)。
    状态: 'new'=可投, 'done'=已沟通过, 'none'=没找到。"""
    # 已经沟通过的会显示"继续沟通"
    e = _first_visible(tab.eles("text:继续沟通", timeout=0.5))
    if e:
        return e, "done"
    for txt in ("立即沟通", "与我沟通", "马上沟通"):
        e = _first_visible(tab.eles(f"text:{txt}", timeout=0.5))
        if e:
            return e, "new"
    # 兜底：按类名找
    e = _first_visible(tab.eles(".btn-startchat", timeout=0.5)) or _first_visible(tab.eles(".op-btn-chat", timeout=0.5))
    if e:
        return e, "new"
    return None, "none"


def chat_input(tab):
    """找聊天输入框。"""
    return (tab.ele("tag:textarea", timeout=2)
            or tab.ele("@contenteditable=true", timeout=0.5)
            or tab.ele(".input-area", timeout=0.5))


LIMIT_HINTS = ("次数已达上限", "次数已用完", "沟通次数已达", "打招呼次数已达",
               "今日剩余沟通", "今日沟通次数", "沟通机会已用", "今日已达上限")


def detect_limit(tab):
    """点完后是否弹出'今日打招呼次数已达上限'之类。"""
    try:
        html = tab.html or ""
    except Exception:
        return False
    return any(h in html for h in LIMIT_HINTS)


def confirm_applied(tab):
    """点完'立即沟通'后，确认是否真的建立了沟通。"""
    try:
        if "/web/geek/chat" in (tab.url or ""):
            return True
    except Exception:
        pass
    if chat_input(tab):
        return True
    try:
        _, state = find_greet_button(tab)
        return state == "done"
    except Exception:
        return False


def send_custom_greeting(tab, greeting):
    """在聊天框里补发一条自定义招呼语（尽力而为，失败不影响投递成功判定）。"""
    if not greeting:
        return
    try:
        ta = chat_input(tab)
        if not ta:
            return
        ta.input(greeting)
        time.sleep(random.uniform(0.6, 1.3))
        btn = (_first_visible(tab.eles("text:发送", timeout=0.6))
               or _first_visible(tab.eles("text:发 送", timeout=0.4)))
        if btn:
            btn.click()
        else:
            ta.input("\n")  # 回车发送兜底
        time.sleep(random.uniform(0.8, 1.5))
    except Exception:
        pass


def apply_one(page, job, link, greeting, dry_run, guard_brief="", guard_model=None):
    """投一个岗位。返回 (status, detail)。status: ok / done / nobtn / error / unfit。"""
    name = str(job.get("岗位名称", ""))[:20]
    company = str(job.get("公司名称", ""))[:20]
    tab = None
    try:
        tab = page.new_tab(link)
        time.sleep(random.uniform(2.0, 3.5))
        btn, state = find_greet_button(tab)
        if state == "done":
            return "done", f"已沟通过 · {name}@{company}"
        if state == "none" or btn is None:
            return "nobtn", f"没找到沟通按钮 · {name}@{company}"
        # 投前 AI 对口核查：读真实职位描述，抓"会CAD/持证/某专业"这类硬要求
        if guard_brief and guard_model:
            jd = read_jd(tab)
            fit, reason = ai_fit_check(guard_brief, job, guard_model, jd=jd)
            if not fit:
                return "unfit", f"专业不对口(看了描述) · {name}@{company} — {reason}"
        if dry_run:
            return "ok", f"[预演]已定位沟通按钮(未发) · {name}@{company}"
        try:
            btn.scroll.to_see()
            time.sleep(random.uniform(0.4, 0.9))
        except Exception:
            pass
        try:
            btn.click()
        except Exception:
            try:
                btn.click(by_js=True)
            except Exception:
                pass
        time.sleep(random.uniform(2.5, 4.0))
        if detect_limit(tab):
            return "limited", f"BOSS提示今日打招呼次数已达上限 · 停在 {name}@{company}"
        if confirm_applied(tab):
            send_custom_greeting(tab, greeting)
            return "ok", f"已投 · {name}@{company}"
        return "error", f"点了沟通但未确认成功 · {name}@{company}"
    except Exception as e:
        # 即便过程抛错，也再确认一次——很多时候其实已经投成功了
        try:
            if tab is not None and confirm_applied(tab):
                return "ok", f"已投(过程有小异常但成功) · {name}@{company}"
        except Exception:
            pass
        return "error", f"出错({str(e)[:40]}) · {name}@{company}"
    finally:
        if tab is not None:
            try:
                time.sleep(random.uniform(0.5, 1.2))
                tab.close()
            except Exception:
                pass


def main():
    cfg = load_config()
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="预演：只定位按钮不真发")
    ap.add_argument("--max", type=int, default=cfg["daily_cap"], help="今天最多投多少(含已投)")
    ap.add_argument("--min-grade", default=cfg["min_grade"], help="投递门槛评级")
    ap.add_argument("--delay-min", type=float, default=cfg["delay_min"])
    ap.add_argument("--delay-max", type=float, default=cfg["delay_max"])
    ap.add_argument("--port", type=int, default=9222)
    ap.add_argument("--table", default=None, help="指定表路径，默认自动找最新精排表")
    ap.add_argument("--monitor-every", type=int, default=cfg.get("monitor_every", 3),
                    help="每投几个跑一次回信监控(0=关闭)")
    ap.add_argument("--no-ai-guard", action="store_true",
                    help="关闭'投递前AI专业对口核查'(默认开启)")
    ap.add_argument("--guard-model", default=None, help="对口核查用的模型(默认跟精排同款)")
    args = ap.parse_args()
    cfg["min_grade"] = args.min_grade

    log("=" * 55)
    log("  BOSS直聘 自动投递" + ("（预演模式·不会真发）" if args.dry_run else "（真投模式）"))
    log("=" * 55)

    table = args.table or latest_table()
    if not table or not os.path.exists(table):
        log("  ❌ 没找到岗位表，请先采集并出一份精排推荐表。")
        return
    log(f"  📄 使用表：{os.path.basename(table)}")

    jobs = read_jobs(table)
    log(f"  表里共 {len(jobs)} 个岗位")

    records = load_records()
    today_done = sum(1 for r in records if r.get("date") == today_str() and r.get("status") == "ok")
    log(f"  今天已投：{today_done} 条；每日上限：{args.max} 条")

    remaining = max(args.max - today_done, 0)
    if remaining <= 0 and not args.dry_run:
        log("  ✅ 今天已达每日上限，自动停止（这是保号措施）。")
        return

    cands, skipped = build_candidates(jobs, records, cfg)
    log(f"  候选 {len(cands)} 个 | 跳过：" +
        " ".join(f"{k}{v}" for k, v in skipped.items() if v))
    if not cands:
        log("  没有可投的新岗位（可能都投过了，或门槛/避坑筛掉了）。")
        return

    # 投递前 AI 对口核查准备（核查移到每个岗位打开详情页后做，能读到真实职位描述）
    guard_brief, guard_model = "", None
    if not args.no_ai_guard:
        guard_brief = load_profile_brief()
        try:
            from local_ai_matcher import DEFAULT_MODEL as _DM, check_ollama
            guard_model = args.guard_model or _DM
            if not guard_brief:
                log("  （没填'我的资料'，跳过AI对口核查；建议先填画像）")
                guard_model = None
            elif not check_ollama(guard_model):
                log("  （本地大模型不可用，跳过AI对口核查，仅用名称黑名单）")
                guard_model = None
            else:
                log(f"  已开启'投前AI对口核查'(读真实职位描述判断，用 {guard_model})")
        except Exception:
            guard_model = None

    if not args.dry_run:
        cands = cands[:remaining]
    log(f"  本次计划处理 {len(cands)} 个" + ("（预演不限量）" if args.dry_run else f"（剩余额度 {remaining}）"))
    log("-" * 55)

    log("  连接调试 Chrome（端口 %d）……" % args.port)
    try:
        page = connect_chrome(args.port)
    except Exception as e:
        log(f"  ❌ 连不上 Chrome：{e}")
        log("     请先点【②打开网站并登录】启动调试 Chrome 并登录 BOSS。")
        return

    try:
        cur = page.url or ""
    except Exception:
        cur = ""
    if "zhipin.com" not in cur:
        log("  ⚠️ 当前 Chrome 不在 BOSS直聘页面，请确认已登录 www.zhipin.com")

    # 回信监控（每投 N 个跑一次），用同一个 page，不必重连
    monitor_every = max(args.monitor_every, 0)
    reply_mod = None
    if monitor_every and not args.dry_run:
        try:
            import reply_monitor as reply_mod
            log(f"  （每投 {monitor_every} 个会自动检查一次回信）")
        except Exception as e:
            log(f"  （回信监控模块加载失败，跳过：{e}）")
            reply_mod = None

    ok_n = done_n = nobtn_n = err_n = unfit_n = 0
    limited = False
    for i, (job, jid, link) in enumerate(cands, 1):
        log(f"  [{i}/{len(cands)}] 处理：{str(job.get('岗位名称',''))[:24]} @ {str(job.get('公司名称',''))[:20]}")
        status, detail = apply_one(page, job, link, cfg["greeting"], args.dry_run,
                                   guard_brief=guard_brief, guard_model=guard_model)
        log(f"      {detail}")
        if status == "limited":
            limited = True
            log("  🛑 BOSS 今日打招呼次数已用完，自动停止（明天再投，这是保号）。")
            break
        if status == "ok":
            ok_n += 1
        elif status == "done":
            done_n += 1
        elif status == "nobtn":
            nobtn_n += 1
        elif status == "unfit":
            unfit_n += 1
        else:
            err_n += 1

        # 每投 N 个成功的，跑一次回信监控
        if reply_mod and status == "ok" and ok_n > 0 and ok_n % monitor_every == 0:
            log(f"  —— 已投 {ok_n} 个，顺手查一次回信 ——")
            try:
                reply_mod.run_check(page, header=False)
            except Exception as e:
                log(f"  （回信检查出错，跳过：{e}）")

        # 记录（预演不写"已投"，但写一条预演记录便于排查；真投才占额度）
        if not args.dry_run and status in ("ok", "done"):
            records.append({
                "job_id": jid, "date": today_str(),
                "time": dt.datetime.now().strftime("%H:%M:%S"),
                "岗位名称": job.get("岗位名称", ""), "公司名称": job.get("公司名称", ""),
                "薪资范围": job.get("薪资范围", ""), "岗位链接": link,
                "status": "ok" if status == "ok" else "done",
            })
            save_records(records)

        # 随机间隔（最后一个不用等）
        if i < len(cands):
            wait = random.uniform(args.delay_min, args.delay_max)
            if args.dry_run:
                wait = random.uniform(0.5, 1.5)  # 预演不用真等
            log(f"      （等 {wait:.0f} 秒，模拟真人）")
            time.sleep(wait)

    # 收尾再查一次回信
    if reply_mod and ok_n > 0:
        log("  —— 收尾：最后查一次回信 ——")
        try:
            reply_mod.run_check(page, header=False)
        except Exception:
            pass

    log("-" * 55)
    log(f"  完成：成功投 {ok_n} | 已沟通过 {done_n} | 专业不对口跳过 {unfit_n} | 没找到按钮 {nobtn_n} | 出错 {err_n}"
        + ("（因BOSS限额提前停止）" if limited else ""))
    if not args.dry_run:
        log(f"  今天累计已投：{today_done + ok_n} / {args.max}")
    if nobtn_n > 0:
        log("  ⚠️ 有岗位没找到'立即沟通'按钮——可能页面结构变了或岗位已关闭，告诉我我来调按钮定位。")
    log("=" * 55)


if __name__ == "__main__":
    main()
