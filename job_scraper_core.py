import datetime as dt
import json
import os
import random
import re
import sys
import time
from urllib.parse import quote

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def configure_console_encoding():
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass


def safe_filename(name):
    name = re.sub(r'[\\/:*?"<>|]+', "_", str(name or "")).strip()
    return name or "招聘岗位采集.xlsx"


def unique_output_path(output_filename):
    output_path = output_filename
    if not os.path.isabs(output_path):
        output_path = os.path.join(BASE_DIR, output_path)

    root, ext = os.path.splitext(output_path)
    if not ext:
        ext = ".xlsx"
        output_path = root + ext

    if not os.path.exists(output_path):
        return output_path

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{root}_{stamp}{ext}"


def split_keywords(raw):
    if not raw:
        return []
    result = []
    seen = set()
    for item in raw.replace("，", ",").split(","):
        keyword = item.strip()
        if keyword and keyword not in seen:
            seen.add(keyword)
            result.append(keyword)
    return result


def load_city_code(city_name):
    path = os.path.join(BASE_DIR, "city_dict.json")
    if not os.path.exists(path) or not city_name:
        return ""
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return ""

    if city_name in data:
        return str(data[city_name])
    for name, code in data.items():
        if city_name in name or name in city_name:
            return str(code)
    return ""


def connect_chrome(debug_port=9222):
    from DrissionPage import ChromiumOptions, ChromiumPage

    try:
        co = ChromiumOptions()
        co.set_local_port(debug_port)
        page = ChromiumPage(co)
        print(f"  已连接到现有 Chrome（调试端口 {debug_port}）")
        return page
    except Exception:
        page = ChromiumPage()
        print("  已启动新的 Chrome 浏览器")
        return page


def build_search_url(config, keyword, city):
    template = config.get("search_url_template") or config["home_url"]
    city_code = load_city_code(city)
    return template.format(
        keyword=quote(keyword or ""),
        city=quote(city or ""),
        city_code=city_code,
    )


_PACE_FACTOR = 1.0


def set_pace(fast=False):
    """快速模式整体缩短防爬等待。"""
    global _PACE_FACTOR
    _PACE_FACTOR = 0.35 if fast else 1.0


def human_pause(min_seconds=0.8, max_seconds=1.8):
    time.sleep(round(random.uniform(min_seconds, max_seconds) * _PACE_FACTOR, 2))


def normalize_text(text):
    return re.sub(r"\s+", " ", str(text or "")).strip()


def extract_salary(text):
    patterns = [
        r"\d+(?:\.\d+)?-\d+(?:\.\d+)?[Kk千]",
        r"\d+(?:\.\d+)?-\d+(?:\.\d+)?万",
        r"\d+(?:\.\d+)?[Kk千]-\d+(?:\.\d+)?[Kk千]",
        r"\d+(?:\.\d+)?万-\d+(?:\.\d+)?万",
        r"\d+(?:\.\d+)?-\d+(?:\.\d+)?万/年",
        r"\d+(?:\.\d+)?-\d+(?:\.\d+)?千/月",
        r"面议|薪资面议",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(0)
    return ""


def extract_company(lines, title):
    company_words = ("公司", "集团", "科技", "网络", "信息", "软件", "有限", "股份", "中心", "事务所")
    for line in lines:
        item = normalize_text(line)
        if item == title or len(item) > 50:
            continue
        if any(word in item for word in company_words):
            return item
    return ""


def extract_requirement(text, words):
    for word in words:
        if word in text:
            return word
    return ""


def extract_active(text):
    """从卡片文本提取 HR/岗位『活跃近况』，用于后续滤僵尸号/已招满未下架。
    例:刚刚活跃、3日内活跃、本周活跃、本月活跃、5个月前活跃、已下线。"""
    s = normalize_text(text)
    m = re.search(
        r"(刚刚活跃|今日活跃|今天活跃|当前在线|在线|本周活跃|本月活跃|"
        r"\d+\s*日内活跃|\d+\s*天内活跃|\d+\s*周前活跃|\d+\s*个?月(?:前|内)活跃|"
        r"半年前活跃|\d{4}年.{0,6}活跃|已下线|停止招聘|已招满)",
        s,
    )
    return m.group(0) if m else ""


def parse_candidate(platform_name, candidate, keyword, city):
    title = normalize_text(candidate.get("title"))
    text = candidate.get("text") or ""
    lines = [normalize_text(x) for x in re.split(r"[\r\n]+", text) if normalize_text(x)]
    flat = normalize_text(text)

    if not title and lines:
        title = lines[0]
    if len(title) > 80:
        title = title[:80]

    company = extract_company(lines, title)
    salary = extract_salary(flat)
    experience = extract_requirement(flat, ["经验不限", "无需经验", "在校/应届", "应届生", "1年以内", "1-3年", "3-5年", "5-10年", "10年以上"])
    degree = extract_requirement(flat, ["学历不限", "初中及以下", "中专/中技", "高中", "大专", "本科", "硕士", "博士"])

    location = city if city and city in flat else ""
    if not location:
        for line in lines:
            if len(line) <= 20 and re.search(r"(区|县|市|镇|路|街|园|新区|开发区)", line):
                location = line
                break

    return {
        "平台": platform_name,
        "搜索城市": city,
        "搜索关键词": keyword,
        "岗位名称": title,
        "薪资范围": salary,
        "公司名称": company,
        "工作地点": location,
        "经验要求": experience,
        "学历要求": degree,
        "岗位链接": candidate.get("link", ""),
        "活跃度": extract_active(flat),
        "原始摘要": flat[:500],
        "采集时间": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def extract_visible_candidates(page, config):
    patterns = config.get("link_patterns", [])
    script = f"""
    (() => {{
      const linkPatterns = {json.dumps(patterns, ensure_ascii=False)};
      const badText = /(登录|注册|首页|简历|隐私|协议|帮助|下载|App|APP|广告|举报)/;
      const text = (el) => (el.innerText || el.textContent || '')
        .replace(/[ \\t\\u00a0]+/g, ' ')
        .replace(/\\n\\s+/g, '\\n')
        .trim();
      const abs = (href) => {{
        try {{ return new URL(href || '', location.href).href; }}
        catch (e) {{ return href || ''; }}
      }};
      const candidates = [];
      for (const a of Array.from(document.querySelectorAll('a[href]'))) {{
        const href = abs(a.getAttribute('href'));
        const title = text(a);
        if (!href || !title || title.length > 90 || badText.test(title)) continue;
        let score = 0;
        if (linkPatterns.some((p) => href.includes(p))) score += 5;
        if (/job|jobs|position|职位|招聘/i.test(href)) score += 2;
        if (/岗位|工程师|专员|经理|主管|助理|顾问|运营|开发|测试|运维|销售|客服|设计|产品/.test(title)) score += 2;
        if (score < 5) continue;

        let node = a;
        let best = a;
        let bestText = text(a);
        for (let i = 0; i < 6 && node.parentElement; i++) {{
          node = node.parentElement;
          const nodeText = text(node);
          const className = String(node.className || '');
          if (nodeText.length > bestText.length && nodeText.length < 1400) {{
            best = node;
            bestText = nodeText;
          }}
          if (/job|职位|card|item|list|position/i.test(className) && nodeText.length > title.length) {{
            best = node;
            bestText = nodeText;
            break;
          }}
        }}
        candidates.push({{ title, link: href, text: bestText.slice(0, 1200) }});
      }}
      const seen = new Set();
      return candidates.filter((item) => {{
        const key = item.link + '|' + item.title;
        if (seen.has(key)) return false;
        seen.add(key);
        return true;
      }}).slice(0, 120);
    }})();
    """
    try:
        return page.run_js(script) or []
    except Exception as e:
        print(f"  ⚠️ 页面提取失败：{e}")
        return []


def scroll_page(page):
    try:
        page.run_js("window.scrollBy({top: Math.max(500, window.innerHeight * 0.9), behavior: 'smooth'});")
    except Exception:
        pass


def click_next_page(page):
    script = """
    (() => {
      const items = Array.from(document.querySelectorAll('a,button'));
      const next = items.find((el) => {
        const text = (el.innerText || el.textContent || '').trim();
        const disabled = el.disabled || /disabled|disable/.test(String(el.className || ''));
        return !disabled && /下一页|下页|Next|next/i.test(text);
      });
      if (!next) return false;
      next.click();
      return true;
    })();
    """
    try:
        return bool(page.run_js(script))
    except Exception:
        return False


def collect_jobs(page, config, keyword, city, pages=1, scrolls=4):
    collected = []
    seen = set()

    for page_no in range(1, pages + 1):
        print(f"  正在采集第 {page_no}/{pages} 页可见岗位...")
        for scroll_no in range(max(1, scrolls)):
            candidates = extract_visible_candidates(page, config)
            for candidate in candidates:
                row = parse_candidate(config["name"], candidate, keyword, city)
                key = (row["岗位链接"], row["岗位名称"], row["公司名称"])
                if row["岗位名称"] and key not in seen:
                    seen.add(key)
                    collected.append(row)
            print(f"    滚动 {scroll_no + 1}/{scrolls}，累计 {len(collected)} 条")
            scroll_page(page)
            human_pause()

        if page_no < pages:
            if click_next_page(page):
                human_pause(2.0, 3.5)
            else:
                print("  未找到可点击的下一页，后续页停止。")
                break

    return collected


# ──────────────────────────────────────────────────────────────────────────
# 接口监听采集（首选方案）：像 BOSS 那样拦截平台真实发出的搜索接口，
# 直接读结构化 JSON，比扫页面链接稳得多。各平台字段名不同，这里做统一映射。
# ──────────────────────────────────────────────────────────────────────────

# 不同平台同一含义的字段名可能不一样，按优先级依次尝试
_FIELD_KEYS = {
    "岗位名称": ["jobName", "positionName", "job_name", "name", "title", "jobTitle",
              "recruitPositionName", "position_name"],
    "薪资范围": ["salaryDesc", "salary", "providesalary_text", "provideSalaryString",
              "salary_text", "salaryReal", "salary60", "salaryString", "salaryRange",
              "wage", "salaryDescription"],
    "公司名称": ["companyName", "brandName", "compName", "fullCompanyName", "company_name",
              "companyFullName", "company", "comp_name", "companyAbbrName"],
    "工作地点": ["cityName", "jobAreaString", "jobArea", "workCity", "city", "dq",
              "areaDistrict", "district", "workAddress", "jobCity"],
    "经验要求": ["jobExperience", "workYearString", "workYear", "requireWorkYears",
              "workyears", "workingExp", "workYearDes", "experience", "workYearCode"],
    "学历要求": ["jobDegree", "eduLevelName", "eduLevel", "education", "eduName",
              "degree", "eduLevelString", "requireEduLevel", "eduRequire", "degreeString"],
    "岗位链接": ["jobHref", "positionUrl", "jobUrl", "positionLink", "url", "link",
              "shareUrl", "detailUrl"],
    "活跃度": ["activeTimeDesc", "lastActiveTime", "activeTime", "lastLoginTime",
             "onlineStatus", "activeTag", "freshStatus", "lastModifyTime"],
}
# 用于从 id 拼链接的常见 id 字段
_ID_KEYS = ["jobId", "positionId", "encryptJobId", "jobid", "positionid", "id",
            "jobIdStr", "showId"]
# 判断一个数组是不是“岗位列表”，元素 dict 至少含一个岗位名字段
_NAME_KEYS = _FIELD_KEYS["岗位名称"]


def _stringify(value):
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, dict):
        for k in ("name", "text", "value", "label", "string", "desc"):
            if value.get(k):
                return str(value[k]).strip()
        return ""
    if isinstance(value, list):
        parts = [_stringify(x) for x in value]
        return " ".join(p for p in parts if p)[:80]
    return ""


def _pick(record, keys):
    for key in keys:
        if key in record:
            text = _stringify(record[key])
            if text:
                return text
    return ""


def _flatten_record(record):
    """把岗位 dict 自身字段与其一层子 dict 字段合并（顶层优先），
    兼容“岗位信息和公司信息分置在 job/comp 子对象里”的结构（如猎聘）。"""
    if not isinstance(record, dict):
        return {}
    flat = {}
    for value in record.values():
        if isinstance(value, dict):
            for key, sub in value.items():
                flat.setdefault(key, sub)
    for key, value in record.items():
        flat[key] = value
    return flat


# 明确的岗位名字段（出现即可判定为岗位，区别于通用的 name）
_STRONG_NAME_KEYS = ["jobName", "positionName", "recruitPositionName", "job_name",
                     "jobTitle", "position_name"]


def _is_job_record(record):
    """判断一条记录是否像“岗位”，避免把国家/城市/行业等字典数组误当岗位。
    规则：① 带明确岗位名字段；或 ② 有通用名 + 薪资或公司信号。"""
    flat = _flatten_record(record)
    if any(_stringify(flat.get(k)) for k in _STRONG_NAME_KEYS):
        return True
    has_name = any(_stringify(flat.get(k)) for k in _NAME_KEYS)
    has_salary = any(_stringify(flat.get(k)) for k in _FIELD_KEYS["薪资范围"])
    has_company = any(_stringify(flat.get(k)) for k in _FIELD_KEYS["公司名称"])
    return has_name and (has_salary or has_company)


def _find_job_records(obj):
    """递归找出 JSON 里最像“岗位列表”的那个数组。"""
    best = []
    if isinstance(obj, list):
        dicts = [x for x in obj if isinstance(x, dict)]
        if dicts:
            jobish = sum(1 for d in dicts if _is_job_record(d))
            if jobish and jobish >= max(1, len(dicts) // 2):
                best = dicts
        for item in obj:
            cand = _find_job_records(item)
            if len(cand) > len(best):
                best = cand
    elif isinstance(obj, dict):
        for value in obj.values():
            cand = _find_job_records(value)
            if len(cand) > len(best):
                best = cand
    return best


def _build_link(config, record):
    link = _pick(record, _FIELD_KEYS["岗位链接"])
    if link:
        if link.startswith("//"):
            return "https:" + link
        if link.startswith("http"):
            return link
        if link.startswith("/"):
            home = config.get("home_url", "").rstrip("/")
            return home + link
        return link
    template = config.get("detail_url_template")
    if template:
        for key in _ID_KEYS:
            if record.get(key) not in (None, "", 0):
                return template.format(id=record[key])
    return ""


def _map_record(config, record, keyword, city):
    record = _flatten_record(record)
    location = _pick(record, _FIELD_KEYS["工作地点"]) or city
    return {
        "平台": config["name"],
        "搜索城市": city,
        "搜索关键词": keyword,
        "岗位名称": _pick(record, _FIELD_KEYS["岗位名称"]),
        "薪资范围": _pick(record, _FIELD_KEYS["薪资范围"]),
        "公司名称": _pick(record, _FIELD_KEYS["公司名称"]),
        "工作地点": location,
        "经验要求": _pick(record, _FIELD_KEYS["经验要求"]),
        "学历要求": _pick(record, _FIELD_KEYS["学历要求"]),
        "岗位链接": _build_link(config, record),
        "活跃度": _pick(record, _FIELD_KEYS.get("活跃度", [])),
        "原始摘要": "",
        "采集时间": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _wait_for_records(page, attempts=8, timeout=8):
    """等待并解析监听到的接口响应，返回岗位 dict 列表；始终没拿到返回 None。"""
    for _ in range(max(1, attempts)):
        try:
            packet = page.listen.wait(timeout=timeout)
        except Exception:
            return None
        if not packet or not getattr(packet, "response", None):
            return None
        body = packet.response.body
        if not body:
            continue
        if isinstance(body, (bytes, bytearray)):
            try:
                body = body.decode("utf-8", "replace")
            except Exception:
                continue
        if isinstance(body, str):
            try:
                body = json.loads(body)
            except Exception:
                continue
        records = _find_job_records(body)
        if records:
            return records
    return None


def collect_jobs_via_api(page, config, keyword, city, pages=1, scrolls=4, manual=False):
    """首选采集路径：监听平台搜索接口。返回 None 表示没捕获到，交给调用方退回页面采集。"""
    patterns = config.get("api_patterns")
    if not patterns:
        return None

    url = build_search_url(config, keyword, city)
    collected = []
    seen = set()
    captured = False

    for page_no in range(1, pages + 1):
        try:
            page.listen.start(patterns)
        except Exception as e:
            print(f"  ⚠️ 无法启动接口监听：{e}")
            return None

        # 触发一次接口请求：第一页（重新）导航/刷新，后续页翻页或下拉
        try:
            if page_no == 1:
                if manual:
                    page.refresh()          # 保留用户手动搜索的结果，仅重新触发接口
                else:
                    page.get(url)
            else:
                if not click_next_page(page):
                    try:
                        page.scroll.to_bottom()
                    except Exception:
                        pass
        except Exception as e:
            print(f"  ⚠️ 翻页/刷新失败：{e}")

        records = _wait_for_records(page, attempts=8, timeout=8)
        try:
            page.listen.stop()
        except Exception:
            pass

        if records is None:
            if page_no == 1:
                return None             # 第一页就没接口 -> 让调用方走页面采集兜底
            break                        # 后续页没数据 -> 当作翻完了

        captured = True
        new_count = 0
        for record in records:
            row = _map_record(config, record, keyword, city)
            key = (row["岗位链接"], row["岗位名称"], row["公司名称"])
            if row["岗位名称"] and key not in seen:
                seen.add(key)
                collected.append(row)
                new_count += 1
        print(f"  ✅ 接口第 {page_no}/{pages} 页：本页 {len(records)} 条，新增 {new_count}，累计 {len(collected)}")
        if new_count == 0:
            break                        # 没新增，多半翻到重复页了
        human_pause()

    return collected if captured else None


def save_jobs_excel(rows, output_filename):
    if not rows:
        return None, 0

    output_path = unique_output_path(output_filename)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    columns = [
        "平台", "搜索城市", "搜索关键词", "岗位名称", "薪资范围", "公司名称",
        "工作地点", "经验要求", "学历要求", "岗位链接", "活跃度", "原始摘要", "采集时间",
    ]
    df = pd.DataFrame(rows)
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    df = df[columns]
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.border = border
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
            elif cell.column in (10, 11):
                cell.alignment = left
            else:
                cell.alignment = center

    widths = {
        "A": 12, "B": 12, "C": 18, "D": 28, "E": 14, "F": 28,
        "G": 18, "H": 12, "I": 12, "J": 44, "K": 60, "L": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    wb.save(output_path)
    return output_path, len(df)
