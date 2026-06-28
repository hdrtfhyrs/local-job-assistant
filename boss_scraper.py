"""
====================================================
  BOSS直聘 景德镇岗位爬虫 - DrissionPage版 v7.0
  新增：无极下拉滚动支持、详情页两阶段采集防风控
====================================================

使用方法：
  1. 在Chrome里打开并登录BOSS直聘 (www.zhipin.com)
  2. 在PyCharm里运行本脚本
  3. 脚本自动接管Chrome，先抓取列表，后针对高分岗位抓取详情
"""

import argparse
import datetime as dt
import functools
import json
import os
import random
import re
import sys
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ====================== 配置区 ======================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_CITY = "景德镇"
DEFAULT_CITY_CODE = "101240800"
DEFAULT_MAX_PAGES = 5
DEFAULT_MAX_DETAIL_COUNT = 30
DEFAULT_OUTPUT_FILENAME = "IT过渡岗位推荐.xlsx"

KEYWORD_CONFIG = [
    {"keyword": "软件实施工程师",  "score": 93, "tier": "S",
     "reason": "本地金蝶/用友/HIS代理商长期招实施，对应届零门槛"},
    {"keyword": "技术支持工程师",  "score": 90, "tier": "S",
     "reason": "陶瓷厂/化工厂/航空配套企业IT支持，只需懂装机/网络"},
    {"keyword": "ERP实施顾问",     "score": 88, "tier": "S",
     "reason": "景德镇工厂大规模上ERP，需驻场培训维护人员"},
    {"keyword": "网络管理员",      "score": 85, "tier": "A",
     "reason": "学校/政府/企业内网管，稳定可备考公务员"},
    {"keyword": "IT运维工程师",    "score": 83, "tier": "A",
     "reason": "昌飞/化工厂/大型制造企业信息中心"},
    {"keyword": "计算机专业教师",  "score": 82, "tier": "A",
     "reason": "中职/高职/培训机构，软件工程本科满足"},
    {"keyword": "数据分析助理",    "score": 80, "tier": "A",
     "reason": "陶瓷电商企业需要能处理Excel/数据的理工科生"},
    {"keyword": "电商运营助理",    "score": 78, "tier": "A",
     "reason": "景德镇陶瓷直播电商特色产业，入职门槛低"},
    {"keyword": "信息管理员",      "score": 77, "tier": "A",
     "reason": "陶瓷/化工企业信息化管理岗，环境稳定"},
    {"keyword": "系统运维",        "score": 75, "tier": "A",
     "reason": "服务器/数据库运维，可积累Linux实操经验"},
    {"keyword": "数字化专员",      "score": 74, "tier": "A",
     "reason": "景德镇制造业数字化转型产生的新型岗位"},
    {"keyword": "测试工程师",      "score": 70, "tier": "B",
     "reason": "本地较少，主要在政务/医疗软件公司"},
    {"keyword": "前端开发工程师",  "score": 62, "tier": "B",
     "reason": "景德镇需求稀少，无项目经验成功率低"},
    {"keyword": "Java开发工程师",  "score": 55, "tier": "C",
     "reason": "景德镇几乎无纯Java开发岗，建议去南昌"},
]

KEYWORD_CONFIG.sort(key=lambda x: x["score"], reverse=True)
KEYWORDS     = [c["keyword"] for c in KEYWORD_CONFIG]
KEYWORD_META = {c["keyword"]: c for c in KEYWORD_CONFIG}

# ====================== 岗位相关性匹配 ======================
# 平台在某城市搜索结果不足时，会用其它“推荐岗位”凑数。下面用于判断
# 返回的岗位名称是否真的与搜索关键词相关，过滤掉牛头不对马嘴的结果。

# 不携带专业含义的通用后缀/修饰词，用于从关键词中提取“核心词干”
_GENERIC_TOKENS = ["工程师", "专员", "助理", "顾问", "教师", "实习生", "实习",
                   "开发", "高级", "初级", "中级", "资深", "员"]

# 关键词 -> 额外可接受的岗位名包含词（提高召回，避免漏掉同义岗位）
MATCH_ALIASES = {
    "软件实施工程师": ["实施", "软件实施", "ERP实施", "上线实施", "驻场实施"],
    "技术支持工程师": ["技术支持", "售后技术", "IT支持", "桌面运维", "运维支持", "售后工程"],
    "ERP实施顾问":   ["ERP", "实施", "用友", "金蝶", "实施顾问"],
    "网络管理员":     ["网络管理", "网管", "网络运维", "弱电", "综合布线"],
    "IT运维工程师":   ["运维", "IT运维", "系统运维", "网络运维", "桌面运维", "IT工程师"],
    "计算机专业教师": ["计算机教师", "信息技术教师", "计算机老师", "专业教师", "实训", "计算机讲师"],
    "数据分析助理":   ["数据分析", "数据专员", "数据运营", "BI", "数据处理"],
    "电商运营助理":   ["电商运营", "运营助理", "店铺运营", "电商专员", "直播运营", "网店运营"],
    "信息管理员":     ["信息管理", "信息化", "信息中心", "信息专员", "信息主管"],
    "系统运维":       ["运维", "系统运维", "服务器运维", "Linux", "数据库运维"],
    "数字化专员":     ["数字化", "信息化", "数智化", "数字化转型"],
    "测试工程师":     ["测试", "软件测试", "测试工程", "QA", "自动化测试"],
    "前端开发工程师": ["前端", "web前端", "h5", "网页开发"],
    "Java开发工程师": ["java", "后端", "j2ee"],
}


def _keyword_stem(keyword):
    stem = keyword or ""
    for token in _GENERIC_TOKENS:
        stem = stem.replace(token, "")
    return stem.strip()


@functools.lru_cache(maxsize=256)
def keyword_match_terms(keyword):
    """生成一个关键词可接受的岗位名包含词集合（全部小写、长度>=2）。"""
    terms = {keyword or ""}
    stem = _keyword_stem(keyword)
    if len(stem) >= 2:
        terms.add(stem)
    terms.update(MATCH_ALIASES.get(keyword, []))
    return frozenset(t.lower() for t in terms if t and len(t) >= 2)


def job_matches_keyword(job_name, keyword):
    """岗位名称是否与搜索关键词相关。"""
    name = (job_name or "").lower()
    if not name:
        return False
    return any(term in name for term in keyword_match_terms(keyword))


def job_in_city(job, city_code, city_name):
    """岗位是否在目标城市。优先用城市代码精确匹配，缺字段时退回城市名，全无则不误删。"""
    code = job.get("city")
    if code not in (None, "", 0):
        if city_code:
            return str(code) == str(city_code)
        return True
    name = (job.get("cityName") or "").strip()
    if name and city_name:
        return city_name in name or name in city_name
    return True  # 没有任何城市信息时，宁可保留也不误删


@functools.lru_cache(maxsize=1)
def load_local_city_dict():
    """读取本地城市代码表，减少网络依赖。"""
    path = os.path.join(BASE_DIR, "city_dict.json")
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {str(k): str(v) for k, v in data.items()}
    except Exception as e:
        print(f"  ⚠️ 读取本地城市表失败: {e}")
        return {}


def get_city_code(city_name, offline=False):
    """通过 Boss直聘 API 获取城市代码"""
    city_name = (city_name or "").strip()
    if not city_name:
        return None

    city_dict = load_local_city_dict()
    if city_name in city_dict:
        return city_dict[city_name]

    for name, code in city_dict.items():
        if city_name in name or name in city_name:
            return code

    if offline:
        return None

    import requests
    try:
        resp = requests.get('https://www.zhipin.com/wapi/zpCommon/data/city.json', 
                            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}, 
                            timeout=5)
        data = resp.json()
        for prov in data.get('zpData', {}).get('cityList', []):
            for city in prov.get('subLevelModelList', []):
                # 支持模糊匹配，例如输入"济南"匹配"济南市"
                if city_name in city['name']:
                    return str(city['code'])
    except Exception as e:
        print(f"  ⚠️ 获取城市代码API失败: {e}")
    return None


def parse_keywords(raw):
    """把中英文逗号分隔的岗位文本转换为去重后的关键词列表。"""
    if not raw:
        return []
    seen = set()
    result = []
    for item in raw.replace("，", ",").split(","):
        keyword = item.strip()
        if keyword and keyword not in seen:
            seen.add(keyword)
            result.append(keyword)
    return result


def build_custom_keyword_meta(keywords):
    return {k: {"score": 100, "tier": "S", "reason": "用户自定义搜索"} for k in keywords}


def safe_filename(name):
    name = re.sub(r'[\\/:*?"<>|]+', "_", name).strip()
    return name or DEFAULT_OUTPUT_FILENAME


def make_output_filename(city, keywords, requested=None):
    if requested:
        return safe_filename(requested)
    if keywords:
        return safe_filename(f"{city}_自定义岗位搜集.xlsx")
    return safe_filename(f"{city}_IT过渡岗位推荐.xlsx")


def unique_output_path(output_filename):
    output_path = output_filename
    if not os.path.isabs(output_path):
        output_path = os.path.join(BASE_DIR, output_path)

    root, ext = os.path.splitext(output_path)
    if not ext:
        ext = ".xlsx"
        output_path = root + ext

    if not os.path.exists(output_path):
        return output_path

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{root}_{stamp}{ext}"


def print_city_matches(keyword="", limit=30):
    city_dict = load_local_city_dict()
    keyword = (keyword or "").strip()
    matches = [
        (name, code) for name, code in city_dict.items()
        if not keyword or keyword in name
    ]
    if not matches:
        print(f"未找到包含「{keyword}」的城市。")
        return
    for name, code in matches[:limit]:
        print(f"{name}: {code}")
    if len(matches) > limit:
        print(f"... 共 {len(matches)} 个匹配，仅显示前 {limit} 个。")


def print_run_config(city, city_code, keywords, max_pages, max_details, output_filename,
                     skip_details, skip_reputation, strict_match=True, fast=False):
    detail_text = "跳过" if skip_details else ("全部" if max_details <= 0 else str(max_details))
    print("\n" + "-" * 55)
    print("  当前运行配置")
    print("-" * 55)
    print(f"  城市：{city} ({city_code or '未匹配到代码'})")
    print(f"  关键词：{', '.join(keywords)}")
    print(f"  每个关键词最多页数：{max_pages}")
    print(f"  详情采集数量：{detail_text}")
    print(f"  公司口碑背调：{'跳过' if skip_reputation else '执行'}")
    print(f"  岗位相关性过滤：{'开启（剔除无关推荐岗位）' if strict_match else '关闭'}")
    print(f"  运行模式：{'⚡ 快速（防爬等待已缩短）' if fast else '常规（防爬保护完整）'}")
    print(f"  输出文件：{output_filename}")
    print("-" * 55)


def build_parser():
    parser = argparse.ArgumentParser(
        description="BOSS直聘岗位采集工具：支持交互式运行，也支持命令行参数批量运行。"
    )
    parser.add_argument("--city", help=f"搜索城市，默认交互输入或 {DEFAULT_CITY}")
    parser.add_argument("--keywords", help="搜索岗位，多个岗位用逗号分隔；不填则使用内置IT过渡岗位")
    parser.add_argument("--max-pages", type=int, default=DEFAULT_MAX_PAGES, help="每个关键词最多抓取页数")
    parser.add_argument("--max-details", type=int, default=DEFAULT_MAX_DETAIL_COUNT, help="最多抓取详情页数量")
    parser.add_argument("--output", help="输出 xlsx 文件名或绝对路径；已存在时会自动追加时间戳")
    parser.add_argument("--skip-details", action="store_true", help="只抓列表，不进入岗位详情页")
    parser.add_argument("--skip-reputation", action="store_true", help="跳过百度口碑背调，速度更快")
    parser.add_argument("--no-strict-match", dest="strict_match", action="store_false",
                        help="关闭岗位相关性过滤（默认开启，会剔除与关键词无关的推荐岗位）")
    parser.set_defaults(strict_match=True)
    parser.add_argument("--fast", action="store_true",
                        help="快速模式：大幅缩短防爬等待，速度快很多，但更容易触发风控/验证码")
    parser.add_argument("--list-cities", nargs="?", const="", metavar="关键字", help="列出本地城市代码，可附带关键字")
    parser.add_argument("--dry-run", action="store_true", help="只解析并打印配置，不启动浏览器")
    parser.add_argument("--yes", action="store_true", help="跳过登录确认和结束暂停，适合已经准备好的调试浏览器")
    parser.add_argument("--debug-port", type=int, default=9222, help="连接已开启 Chrome 的远程调试端口")
    return parser


def configure_console_encoding():
    """避免 Windows GBK 控制台遇到 emoji 或中文时中断程序。"""
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass

# ====================== 运行节奏（防爬等待）======================

class Pace:
    """统一管理防爬等待。--fast 会整体收紧，明显更快但更容易触发风控。"""

    def __init__(self, fast=False):
        self.fast = fast
        self.factor = 0.35 if fast else 1.0      # 随机等待整体缩放
        self.api_timeout = 5 if fast else 8       # 监听 API 的等待上限
        self.retries = 2 if fast else 3           # 单页最多尝试次数
        self.detail_timeout = 3 if fast else 6    # 详情页元素等待

    def nap(self, lo, hi):
        time.sleep(round(random.uniform(lo, hi) * self.factor, 2))


# ====================== 阶段一：列表爬取 ======================

def scrape_all(page, city_code, keywords, keyword_meta, max_pages=DEFAULT_MAX_PAGES,
               strict_match=True, city_name="", pace=None):
    """用DrissionPage监听API，无限下拉遍历抓取"""
    from urllib.parse import quote

    pace = pace or Pace()

    all_jobs = []

    def search_url(keyword):
        # 直接用 URL 锁定“关键词 + 城市”，避免在搜索框打字时被账号定位城市覆盖
        u = "https://www.zhipin.com/web/geek/job?query=" + quote(keyword)
        if city_code:
            u += f"&city={city_code}"
        return u

    for kw_idx, keyword in enumerate(keywords):
        meta = keyword_meta[keyword]
        print("\n" + "=" * 55)
        print("  [{}/{}] 关键词：【{}】  推荐分：{}".format(
            kw_idx + 1, len(keywords), keyword, meta["score"]))
        print("=" * 55)

        kw_jobs = []
        wrong_city_total = 0

        for page_num in range(1, max_pages + 1):
            # 每页开头把抓取页强制设为"活跃"，绕开窗口最小化/被遮挡导致的后台节流
            try:
                page.run_cdp("Page.setWebLifecycleState", state="active")
            except Exception:
                pass
            page.listen.start("wapi/zpgeek/search/joblist.json")

            # ── 触发搜索或下拉 ──────────────────────────────────────
            if page_num == 1:
                # 第一页：直接导航到带关键词与城市的搜索 URL
                page.get(search_url(keyword))
                print("  打开搜索页：{} @ {}".format(keyword, city_name or "默认城市"))
                pace.nap(0.6, 1.0)
            else:
                # 翻页：模拟真人向下滚动（无限下拉加载）
                try:
                    page.scroll.to_bottom()
                    # 增加一些抖动更真实
                    page.scroll.up(300)
                    pace.nap(0.3, 0.5)
                    page.scroll.to_bottom()
                    print("  向下滚动加载第 {} 页...".format(page_num))
                except Exception as e:
                    print("  📌  滚动失败：{}".format(e))
                    page.listen.stop()
                    break

            # ── 等待并重试获取数据 ───────────────────────────────────
            job_list = []
            for attempt in range(pace.retries):
                try:
                    resp = page.listen.wait(timeout=pace.api_timeout)
                    if hasattr(resp, "response") and resp.response:
                        body = resp.response.body
                        if isinstance(body, str):
                            body = json.loads(body)
                        
                        code = body.get("code", -1)
                        if code == 0:
                            job_list = body.get("zpData", {}).get("jobList", [])
                            if job_list:
                                break  # 成功拿到数据，跳出重试循环
                            else:
                                if page.ele('text:没有找到相关', timeout=1) or page.ele('text:修改一下搜索条件', timeout=1):
                                    print("  📌  页面提示无相关职位，真实0结果，停止重试。")
                                    break
                                print("  ⚠️  第{}次请求数据为空，正在重试...".format(attempt + 1))
                        else:
                            print("  ⚠️  第{}次请求API code={}，正在重试...".format(attempt + 1, code))
                    else:
                        print("  ⚠️  第{}次请求无响应体，正在重试...".format(attempt + 1))
                except Exception:
                    print("  ⚠️  第{}次等待API超时，正在重试...".format(attempt + 1))

                # 重试机制：第一页重新导航搜索 URL，其余页再次滚动
                pace.nap(1.0, 2.0)
                if page_num == 1:
                    print("  -> 尝试重新打开搜索页...")
                    page.get(search_url(keyword))
                    pace.nap(0.6, 1.0)
                elif page_num > 1:
                    print("  -> 尝试再次向下滚动...")
                    page.scroll.up(500)
                    pace.nap(0.3, 0.5)
                    page.scroll.to_bottom()

            page.listen.stop()

            if not job_list:
                print("  📌  第{}页多次尝试后仍无数据，可能是末页或被防爬拦截".format(page_num))
                break

            # ── 解析响应数据 ──────────────────────────────────────
            try:
                matched = 0
                skipped_irrelevant = 0
                skipped_wrong_city = 0
                for job in job_list:
                    job_name = job.get("jobName", "")
                    enc_id   = job.get("encryptJobId", "")

                    # 没有有效详情ID的多是推荐占位/失效岗位，链接在平台上打不开
                    if not enc_id:
                        continue
                    # 过滤掉与关键词无关的“推荐”岗位（搜索结果不足时平台会拿其它岗位凑数）
                    if strict_match and not job_matches_keyword(job_name, keyword):
                        skipped_irrelevant += 1
                        continue
                    # 过滤掉非目标城市的岗位（平台常按账号定位城市返回外地岗位）
                    if strict_match and not job_in_city(job, city_code, city_name):
                        skipped_wrong_city += 1
                        continue

                    city_nm = (job.get("cityName") or "").strip()
                    area    = job.get("areaDistrict", "")
                    biz     = job.get("bizDistrict", "") or ""
                    loc     = (city_nm + " " if city_nm else "") + area + (" " + biz if biz else "")
                    kw_jobs.append({
                        "推荐分":    meta["score"],
                        "推荐等级":  meta["tier"],
                        "推荐理由":  meta["reason"],
                        "搜索关键词": keyword,
                        "岗位名称":  job_name,
                        "薪资范围":  job.get("salaryDesc", ""),
                        "公司名称":  job.get("brandName", ""),
                        "公司规模":  job.get("brandScaleName", ""),
                        "融资阶段":  job.get("brandStageName", ""),
                        "工作地点":  loc,
                        "经验要求":  job.get("jobExperience", ""),
                        "学历要求":  job.get("jobDegree", ""),
                        "岗位链接":  "https://www.zhipin.com/job_detail/{}.html".format(enc_id),
                        "岗位详情":  "",  # 为详情页预留
                        "具体地址":  "",  # 为详情页预留
                        "网络口碑评价": "", # 为背调预留
                        "背调风险分": ""   # 为背调预留
                    })
                    matched += 1

                wrong_city_total += skipped_wrong_city
                notes = []
                if skipped_irrelevant:
                    notes.append("滤掉 {} 个不相关".format(skipped_irrelevant))
                if skipped_wrong_city:
                    notes.append("滤掉 {} 个外地".format(skipped_wrong_city))
                note = "（{}）".format("，".join(notes)) if notes else ""
                print("  ✅  第{}页完成，命中 {} 个，本词累计 {} 个{}".format(
                    page_num, matched, len(kw_jobs), note))

                # 整页都不相关，说明已进入平台的“推荐”信息流，没必要再翻下去
                if strict_match and matched == 0:
                    print("  📌  本页无与【{}】相关的岗位，停止抓取该词".format(keyword))
                    break

                if len(job_list) < 15:
                    print("  📌  结果不足一页，已抓完")
                    break
                
                # 智能判断是否还有下一页
                try:
                    has_more = body.get("zpData", {}).get("hasMore", True)
                    if not has_more:
                        print("  📌  API返回已无更多数据(hasMore=False)，已抓完")
                        break
                except Exception:
                    pass

            except Exception as e:
                print("  ❌  解析第{}页失败：{}".format(page_num, e))
                break

            # 模拟人类翻页间隔
            pace.nap(2.5, 5.0)

        print("  🎯  【{}】共 {} 个岗位".format(keyword, len(kw_jobs)))
        if city_name and not kw_jobs and wrong_city_total > 0:
            print("  ⚠️  本词命中的 {} 个岗位都不在【{}】，已全部过滤。".format(wrong_city_total, city_name))
            print("      多半是浏览器/账号定位城市不对，请在网页左上角把城市切到【{}】后重跑。".format(city_name))
        all_jobs.extend(kw_jobs)

        # 关键词间休息（没抓到任何岗位时无需再装作浏览，直接切下一个）
        if kw_idx < len(keywords) - 1 and kw_jobs:
            print("  😴  切换关键词，休息片刻...")
            pace.nap(4, 8)

    return all_jobs


# ====================== 阶段二：详情页采集 ======================

def dedupe_jobs(all_jobs):
    """跨关键词去重：同一岗位可能被多个关键词命中，按推荐分降序后只保留最高的一条。"""
    if not all_jobs:
        return all_jobs
    df = pd.DataFrame(all_jobs)
    df.sort_values(by="推荐分", ascending=False, inplace=True)
    df.drop_duplicates(subset=["岗位名称", "公司名称", "薪资范围"], keep="first", inplace=True)
    return df.to_dict("records")


# 详情页用于判断岗位是否已下线/不存在的提示文案
_OFFLINE_HINTS = ["该职位已经关闭", "该职位已关闭", "职位已下线",
                  "职位不存在", "该职位已下线", "职位已经被删除"]


def scrape_details(page, all_jobs, max_detail_count=DEFAULT_MAX_DETAIL_COUNT, pace=None):
    """提取排名前 N 的岗位详情，采用极高防爬标准"""
    pace = pace or Pace()
    if not all_jobs:
        return all_jobs

    # 去重并按分数排序
    unique_jobs = dedupe_jobs(all_jobs)

    print("\n" + "=" * 55)
    print("  🚀 进入第二阶段：采集详情页数据")
    print(f"  总去重岗位数：{len(unique_jobs)}")
    
    if max_detail_count <= 0:
        target_jobs = unique_jobs
    else:
        target_jobs = unique_jobs[:max_detail_count]
    print(f"  详情采集目标：{len(target_jobs)}/{len(unique_jobs)} 条")
    print("=" * 55)

    for i, job in enumerate(target_jobs):
        url = job["岗位链接"]
        print(f"  [{i+1}/{len(target_jobs)}] 正在抓取: {job['公司名称']} - {job['岗位名称']}")
        
        try:
            page.get(url)

            # 检测“幽灵岗位”：列表里有、但平台详情页已下线/不存在
            page_text = ""
            try:
                page_text = page.html
            except Exception:
                page_text = ""
            if any(hint in page_text for hint in _OFFLINE_HINTS):
                job["岗位详情"] = "⚠️ 职位已下线/不存在（平台已无法打开）"
                job["具体地址"] = ""
                print("  ⚠️ 该岗位已下线，标记后跳过")
                pace.nap(1.5, 3.0)
                continue

            # 等待内容加载
            jd_ele = page.ele('css:.job-sec-text', timeout=pace.detail_timeout)
            addr_ele = page.ele('css:.location-address', timeout=2)

            if jd_ele:
                job["岗位详情"] = jd_ele.text.strip()
            else:
                job["岗位详情"] = "未获取到详情"
                
            if addr_ele:
                job["具体地址"] = addr_ele.text.strip()
            else:
                job["具体地址"] = "未获取到地址"
                
            print("  ✅ 成功获取详情数据")

            # 随机停顿防封禁（真实人类看一个岗位的速度）
            pace.nap(3.5, 6.5)

        except Exception as e:
            print(f"  ❌ 抓取失败：{e}")
            job["岗位详情"] = "抓取失败"
            # 遇到严重错误（可能被封），延长等待
            time.sleep(10)

    # 剩余的岗位（如果有的话）详情为空
    return unique_jobs


# ====================== 阶段三：公司口碑背调 ======================

# 把负面/正面信号按类别归类，便于给出“详细理由”而不是一堆散词。
# 注意：刻意不收录“坑/避雷”等会出现在搜索框回显里的裸词，改用更明确的“坑人/割韭菜”。
NEG_CATEGORIES = {
    "薪资欠薪": ["拖欠", "不发工资", "克扣", "压工资", "欠薪", "工资低", "降薪", "提成不发"],
    "加班严重": ["强制加班", "大小周", "996", "007", "加班严重", "无偿加班", "夜班多"],
    "用工性质": ["外包", "劳务派遣", "派遣", "中介公司"],
    "诚信风险": ["骗子", "毁约", "跑路", "黑心", "坑人", "割韭菜", "骗钱", "传销"],
    "经营稳定": ["裁员", "倒闭", "欠债", "经营异常", "失信", "被执行"],
    "职场环境": ["PUA", "画饼", "压榨", "氛围差", "内卷", "勾心斗角"],
}
POS_CATEGORIES = {
    "作息友好": ["双休", "不加班", "准时下班", "弹性工作", "周末双休"],
    "福利待遇": ["五险一金", "福利好", "包吃", "包住", "年终奖", "节日福利", "餐补"],
    "团队氛围": ["人性化", "氛围好", "有人带", "靠谱", "成长空间", "培训完善"],
}

# 百度反爬/验证页特征
_BAIDU_BLOCK_HINTS = ["百度安全验证", "请输入验证码", "网络不给力", "/static/captcha", "安全验证"]


def _result_snippet(text, word, window=22):
    """截取命中词周围的一小段原文，作为证据。"""
    idx = text.find(word)
    if idx == -1:
        return ""
    a = max(0, idx - window)
    b = min(len(text), idx + len(word) + window)
    snippet = re.sub(r"\s+", " ", text[a:b]).strip()
    return f"…{snippet}…"


def _scan_categories(text, categories):
    """返回 {类别: (命中次数, 命中词列表, 代表性原文片段)}。"""
    found = {}
    for cat, words in categories.items():
        hits, matched, snippet = 0, [], ""
        for word in words:
            c = text.count(word)
            if c:
                hits += c
                matched.append(word)
                if not snippet:
                    snippet = _result_snippet(text, word)
        if hits:
            found[cat] = (hits, matched, snippet)
    return found


def scrape_reputation(page, jobs, pace=None):
    """百度搜索做公司口碑背调：只分析搜索结果区文本，输出带证据的详细评价。"""
    import urllib.parse

    pace = pace or Pace()

    print("\n" + "=" * 55)
    print("  🕵️‍♂️ 进入第三阶段：公司口碑背调 (信息源自百度，仅供参考)")
    print("=" * 55)

    company_cache = {}

    for i, job in enumerate(jobs):
        company = job.get("公司名称", "")
        if not company or "未获取" in company:
            continue

        print(f"  [{i+1}/{len(jobs)}] 正在调查公司: {company}")

        if company in company_cache:
            job["网络口碑评价"] = company_cache[company]["comment"]
            job["背调风险分"] = company_cache[company]["score"]
            print("  ✅ (已读取缓存)")
            continue

        try:
            # 关键：查询词里不放“避雷/坑”等负面词，否则百度会把它们回显进页面造成误判
            query = f'"{company}" 怎么样 待遇 评价 工作体验'
            url = "https://www.baidu.com/s?wd=" + urllib.parse.quote(query)
            page.get(url)
            pace.nap(1.5, 3.0)

            # 只取搜索结果容器的可见文本，排除导航/广告/脚本/查询回显等噪音
            container = page.ele("css:#content_left", timeout=4)
            result_text = container.text if container else ""

            # 识别百度验证/反爬页，避免把验证页当成“无负面=正常”
            if not result_text or any(h in page.html for h in _BAIDU_BLOCK_HINTS):
                risk = "未背调(遇验证)"
                comment = ("ℹ️ 背调中断：百度触发了安全验证或暂无结果，本条未能完成口碑分析。\n"
                           f"　建议手动搜索核实：{query}")
                job["网络口碑评价"] = comment
                job["背调风险分"] = risk
                company_cache[company] = {"comment": comment, "score": risk}
                print(f"  ⚠️ {risk}，跳过")
                pace.nap(2.0, 3.5)
                continue

            neg = _scan_categories(result_text, NEG_CATEGORIES)
            pos = _scan_categories(result_text, POS_CATEGORIES)
            neg_cat, neg_total, pos_cat = len(neg), sum(v[0] for v in neg.values()), len(pos)

            # 风险评级：看“命中几类负面 + 总提及次数”，比单纯计数更稳健
            if neg_cat >= 3 or neg_total >= 8:
                risk = "高危(建议避雷)"
            elif neg_cat >= 2 or neg_total >= 4:
                risk = "中度风险"
            elif neg_cat >= 1:
                risk = "轻微提示"
            elif pos_cat >= 1:
                risk = "口碑良好"
            else:
                risk = "信息较少"

            # 组装“详细理由”：判定 + 命中类别明细 + 1~2 条原文证据
            lines = [f"风险判定：{risk}（依据百度搜索结果，仅供参考）"]
            if neg:
                parts = [
                    f"{cat}×{hits}（{'/'.join(words[:3])}）"
                    for cat, (hits, words, _s) in sorted(neg.items(), key=lambda kv: -kv[1][0])
                ]
                lines.append("负面线索：" + "；".join(parts))
                for e in [s for (_h, _w, s) in neg.values() if s][:2]:
                    lines.append("　原文：" + e)
            if pos:
                parts = [f"{cat}（{'/'.join(words[:2])}）" for cat, (_h, words, _s) in pos.items()]
                lines.append("正面线索：" + "；".join(parts))
            if not neg and not pos:
                lines.append("说明：结果中未出现明显的高频好评或差评关键词，信息量偏少，建议自行核实。")

            comment = "\n".join(lines)
            job["网络口碑评价"] = comment
            job["背调风险分"] = risk
            company_cache[company] = {"comment": comment, "score": risk}

            print(f"  📊 调查结果: {risk}（负面 {neg_cat} 类/{neg_total} 次，正面 {pos_cat} 类）")

            pace.nap(4.0, 7.5)

        except Exception as e:
            print(f"  ❌ 背调失败：{e}")
            job["网络口碑评价"] = "背调失败"
            job["背调风险分"] = "-"
            time.sleep(5)

    return jobs


# ====================== 保存与美化 ======================

def save_and_beautify(all_jobs, output_filename):
    if not all_jobs:
        print("\n❌  没有抓取到任何数据！")
        return None

    df = pd.DataFrame(all_jobs)

    output_path = unique_output_path(output_filename)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    TIER_FILLS = {
        "S": PatternFill("solid", fgColor="FFD700"),
        "A": PatternFill("solid", fgColor="C6EFCE"),
        "B": PatternFill("solid", fgColor="BDD7EE"),
        "C": PatternFill("solid", fgColor="D9D9D9"),
    }
    TIER_FONTS = {
        "S": Font(bold=True, color="7B3F00", size=10),
        "A": Font(bold=True, color="375623", size=10),
        "B": Font(bold=True, color="1F497D", size=10),
        "C": Font(bold=False, color="7F7F7F", size=10),
    }
    header_fill  = PatternFill("solid", fgColor="1F3864")
    header_font  = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
    thin         = Side(style="thin")
    thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    tier_col = None
    for cell in ws[1]:
        if cell.value == "推荐等级":
            tier_col = cell.column
            break

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        tier     = ws.cell(row=row_idx, column=tier_col).value if (tier_col and row_idx > 1) else None
        row_fill = TIER_FILLS.get(tier, PatternFill("solid", fgColor="FFFFFF"))
        row_font = TIER_FONTS.get(tier, Font(size=10))
        for cell in row:
            cell.border = thin_border
            if row_idx == 1:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = center_align
            else:
                cell.fill      = row_fill
                cell.font      = row_font
                # 详情页和链接需要左对齐
                if cell.column_letter in ["C", "M", "N", "O", "P", "Q"]:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

    # 调整列宽
    col_widths = {
        "A": 10, "B": 10, "C": 42, "D": 14,
        "E": 20, "F": 16, "G": 22, "H": 14,
        "I": 14, "J": 18, "K": 12, "L": 12, "M": 40,
        "N": 60, "O": 30, "P": 50, "Q": 16
    }
    for col, w in col_widths.items():
        if col in ws.column_dimensions or True:
            ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    wb.save(output_path)
    return output_path, len(df)


# ====================== 主程序 ======================

def main(argv=None):
    configure_console_encoding()
    args = build_parser().parse_args(argv)
    arg_count = len(sys.argv[1:] if argv is None else argv)
    interactive = (not args.yes) and sys.stdin.isatty() and arg_count == 0

    if args.list_cities is not None:
        print_city_matches(args.list_cities)
        return

    print("\n" + "=" * 55)
    print("  BOSS直聘爬虫 DrissionPage版 v8.0")
    print("  特性: 自定义城市与职业、列表/详情/背调可控、结果防覆盖")
    print("=" * 55)
    print()

    # --- 自定义配置输入 ---
    if args.city:
        city_input = args.city.strip()
    elif not interactive:
        city_input = DEFAULT_CITY
    else:
        print("  [自定义配置] (直接按回车则使用系统默认设置)")
        city_input = input(f"  👉 请输入你要搜索的城市 (例如: 济南, 默认: {DEFAULT_CITY}): ").strip()
        if not city_input:
            city_input = DEFAULT_CITY
        
    city_code = get_city_code(city_input)
    if not city_code:
        print(f"  ⚠️ 未能自动找到城市 '{city_input}' 的代码，搜索页可能默认定位到你当前所在城市。")
    else:
        print(f"  ✅ 城市匹配成功: {city_input} (代码: {city_code})")

    if args.keywords is not None:
        kw_input = args.keywords.strip()
    elif not interactive:
        kw_input = ""
    else:
        kw_input = input("  👉 请输入你要搜索的职业岗位，多个岗位请用逗号分隔 (默认: 内置IT精选岗位): ").strip()
    
    current_keywords = []
    current_keyword_meta = {}
    custom_keywords = parse_keywords(kw_input)
    if custom_keywords:
        # 使用用户自定义岗位
        current_keywords = custom_keywords
        current_keyword_meta = build_custom_keyword_meta(current_keywords)
    else:
        # 使用默认岗位配置
        current_keywords = KEYWORDS
        current_keyword_meta = KEYWORD_META
        print("  ✅ 已加载默认 IT 岗位列表")

    max_pages = max(1, args.max_pages)
    max_details = max(0, args.max_details)
    output_filename = make_output_filename(city_input, custom_keywords, args.output)
    print_run_config(
        city_input,
        city_code,
        current_keywords,
        max_pages,
        max_details,
        output_filename,
        args.skip_details,
        args.skip_reputation,
        args.strict_match,
        args.fast,
    )

    if args.dry_run:
        print("  dry-run 完成：未启动浏览器、未抓取数据。")
        return
    
    print("\n" + "-" * 55)

    from DrissionPage import ChromiumPage, ChromiumOptions

    # 尝试连接已打开的Chrome（端口9222），或启动新Chrome
    try:
        co = ChromiumOptions()
        co.set_local_port(args.debug_port)           # 连接已有Chrome
        page = ChromiumPage(co)
        print(f"  已连接到现有Chrome浏览器（调试端口 {args.debug_port}）")
    except Exception:
        # 没有已开启的Chrome，自动启动一个新的
        try:
            page = ChromiumPage()
            print("  已启动新Chrome浏览器")
        except Exception as e:
            print("  启动Chrome失败：{}".format(e))
            if interactive:
                input("  按回车退出...")
            return

    # 检查是否已登录
    try:
        current = page.url or ""
    except Exception:
        current = ""

    if "zhipin.com" not in current:
        page.get("https://www.zhipin.com/")
        time.sleep(3)

    # 等待登录
    print()
    print("  请在Chrome窗口里确认已登录BOSS直聘（右上角有头像）")
    if not city_code:
        print(f"  ⚠️ 因为未匹配到城市代码，请手动在网页左上角把城市切换为: {city_input}")
    print("  如果还没登录，请先扫码登录")
    print()
    if interactive:
        input("  >>> 确认已登录（且城市正确）后，按【回车】开始抓取：")
    print()
    print("  开始自动抓取……")
    print()

    pace = Pace(fast=args.fast)

    # 新开一个独立标签页专门抓取：在它上面导航/滚动/监听API，
    # 你在原来的页面怎么滚动、点击、切换都不会打断抓取（标签页互不干扰）。
    try:
        scrape_tab = page.new_tab()
        print("  ✅ 已新开一个标签页用于抓取，你原来的页面随便看，别关闭 Chrome 就行。")
    except Exception as e:
        scrape_tab = page
        print(f"  （新建标签页失败，改用当前页抓取：{e}；抓取时尽量别动浏览器）")
    print()

    # 第一阶段：获取列表
    all_jobs = scrape_all(scrape_tab, city_code, current_keywords, current_keyword_meta,
                          max_pages=max_pages, strict_match=args.strict_match,
                          city_name=city_input, pace=pace)

    # 跨关键词去重（任何分支都生效，避免重复岗位）
    before = len(all_jobs)
    all_jobs = dedupe_jobs(all_jobs)
    if before != len(all_jobs):
        print(f"\n  🧹 跨关键词去重：{before} → {len(all_jobs)} 条")

    # 第二阶段：获取详情
    if args.skip_details:
        final_jobs = all_jobs
        print("\n  已按配置跳过详情页采集。")
    else:
        final_jobs = scrape_details(scrape_tab, all_jobs, max_detail_count=max_details, pace=pace)

    # 第三阶段：公司背调
    if args.skip_reputation:
        print("\n  已按配置跳过公司口碑背调。")
    else:
        final_jobs = scrape_reputation(scrape_tab, final_jobs, pace=pace)

    # 抓取结束，关掉这个临时标签页，把界面还给你
    try:
        if scrape_tab is not page:
            scrape_tab.close()
    except Exception:
        pass

    # 保存
    result = save_and_beautify(final_jobs, output_filename)
    if result:
        output_path, count = result
        print()
        print("=" * 55)
        print("  🎉 全部完成！共保存 {} 条精选岗位数据".format(count))
        print("  📁 结果已保存至：{}".format(output_path))
        print("=" * 55)
    
    if interactive:
        input("\n  按回车关闭...")


if __name__ == "__main__":
    main()
