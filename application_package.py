import os
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from job_scraper_core import configure_console_encoding


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOCAL_CHANNELS_PATH = os.path.join(BASE_DIR, "local_niche_channels.json")


PLATFORM_ROWS = [
    {
        "平台": "BOSS直聘",
        "入口": "https://www.zhipin.com/",
        "投递特点": "先沟通再投递/发送简历更常见，适合主动打招呼。",
        "必须准备": "手机号登录、在线简历完整、头像/姓名/求职状态、常用招呼语。",
        "容易卡住": "岗位匹配低无人回复；未登录；城市/期望岗位不清；无法直接发附件。",
        "建议动作": "先筛近期活跃Boss，开场说明普通本科+愿意从实施/支持/运营做起。",
    },
    {
        "平台": "前程无忧",
        "入口": "https://www.51job.com/",
        "投递特点": "传统投递，通常依赖在线简历和附件简历。",
        "必须准备": "注册账号、完善在线简历、上传附件简历、设置求职意向。",
        "容易卡住": "在线简历字段不完整；附件格式不对；职位已关闭；城市筛选错误。",
        "建议动作": "先完善在线简历，再投递；投递后记录公司和岗位，避免重复乱投。",
    },
    {
        "平台": "智联招聘",
        "入口": "https://www.zhaopin.com/",
        "投递特点": "在线简历投递为主，也适合找本地传统企业。",
        "必须准备": "手机号登录、在线简历、求职状态、期望城市/岗位。",
        "容易卡住": "搜索结果多但不匹配；岗位要求经验过高；简历亮点弱。",
        "建议动作": "优先投应届/经验不限/有人带/双休/五险一金关键词岗位。",
    },
    {
        "平台": "猎聘",
        "入口": "https://www.liepin.com/",
        "投递特点": "中高端岗位更多，但也可找技术支持、实施、运营类岗位。",
        "必须准备": "在线简历、期望行业、期望薪资、工作地点。",
        "容易卡住": "岗位要求偏高；猎头沟通不稳定；初级岗位少。",
        "建议动作": "少量投递，不作为主平台；遇到猎头先问岗位基础信息和公司性质。",
    },
    {
        "平台": "企业官网/公众号",
        "入口": "公司官网/公众号/本地招聘公众号",
        "投递特点": "绕开平台推荐算法，适合本地企业、文旅、陶瓷、电商公司。",
        "必须准备": "PDF简历、邮件标题、求职说明、电话/微信。",
        "容易卡住": "官网入口难找；邮箱无人看；岗位信息不完整。",
        "建议动作": "平台看到公司后，再搜公司官网/公众号二次投递；3个工作日后跟进。",
    },
]


CHECKLIST_ROWS = [
    ("基础信息", "姓名、手机号、邮箱、微信号一致，手机号能接短信。"),
    ("求职意向", "城市写景德镇/可接受周边；岗位写实施、技术支持、运营助理、数据助理等。"),
    ("在线简历", "教育经历、项目经历、技能、实习/兼职、证书、求职优势都填完整。"),
    ("附件简历", "准备 PDF 和 DOCX 两版，文件名：姓名_岗位方向_手机号。"),
    ("岗位关键词", "不要只搜软件开发；重点搜实施、技术支持、IT运维、电商运营、数据助理。"),
    ("避坑筛选", "看到单休、大小周、夜班、重体力、无社保、纯销售，先降级或不投。"),
    ("投递记录", "每投一个岗位记录平台、公司、岗位、链接、日期、状态和下次跟进日。"),
    ("跟进动作", "BOSS类平台当天发开场白；传统平台2-3个工作日后再查状态。"),
]


TRACKER_HEADERS = [
    "投递日期", "平台", "公司", "岗位", "城市", "岗位方向", "链接",
    "匹配度(1-5)", "工作时长风险", "体力风险", "前景", "本地特色",
    "是否已投", "是否已沟通", "HR回复", "下一步", "下次跟进日", "备注",
]


SCRIPT_ROWS = [
    {
        "场景": "BOSS开场白-实施/技术支持",
        "话术": "您好，我是软件工程本科毕业生，想找景德镇本地的软件实施/技术支持/信息化相关岗位。虽然开发基础不算很强，但我愿意从客户沟通、系统配置、问题排查、文档记录做起，也能补SQL、网络和办公系统。想了解下这个岗位是否接受应届/初级候选人？",
    },
    {
        "场景": "BOSS开场白-陶瓷电商/运营助理",
        "话术": "您好，我是软件工程本科毕业生，想往景德镇陶瓷电商/运营助理方向发展。熟悉电脑和基础数据处理，愿意学习店铺后台、商品上架、数据报表和售后流程。想问下这个岗位是否有人带，以及主要工作内容偏运营还是销售？",
    },
    {
        "场景": "邮件标题",
        "话术": "应聘【岗位名称】- 软件工程本科 - 姓名 - 手机号",
    },
    {
        "场景": "邮件正文",
        "话术": "您好，我在招聘平台看到贵公司的【岗位名称】。我本科软件工程毕业，希望在景德镇长期发展，优先考虑实施、技术支持、运营/数据助理等能持续学习的岗位。附件是我的简历，期待有机会沟通，谢谢。",
    },
    {
        "场景": "面试前必问",
        "话术": "请问这个岗位是否双休/大小周？是否需要夜班或长期出差？试用期是否缴纳社保？有没有新人培训或导师？日常工作中销售指标占比高吗？",
    },
]


RESUME_MARKDOWN = """# 简历投递版模板

> 使用方法：把【】里的内容替换成你的真实信息，导出 PDF 后作为附件简历。在线简历也按这个结构填写。

## 基本信息

- 姓名：【你的姓名】
- 手机：【手机号】
- 邮箱：【邮箱】
- 求职城市：【景德镇/可接受周边】
- 求职方向：【软件实施/技术支持/IT运维/电商运营助理/数据助理】

## 求职优势

- 软件工程本科，有计算机基础，愿意从实施、支持、运维、运营/数据助理等岗位踏实积累。
- 能使用 Word、Excel、PPT，愿意学习 SQL、网络基础、店铺后台、工单系统和业务流程。
- 接受从基础工作做起，重视沟通、记录、复盘和稳定成长。

## 教育经历

【学校名称】｜软件工程｜本科｜【毕业年份】

相关课程：数据库、软件工程、计算机网络、操作系统、Java/Python/Web基础、软件测试。

## 项目经历

### 【课程项目/毕业设计名称】

- 项目简介：【用一句话写系统是做什么的】
- 你负责：【需求整理/页面/数据库/测试/文档/部署】
- 用到技术：【Java/Python/HTML/CSS/SQL/Excel等】
- 可迁移能力：需求理解、问题排查、文档记录、和同学协作完成交付。

## 技能

- 办公与数据：Excel基础、表格整理、数据筛选、简单报表。
- 计算机基础：Windows常见问题处理、网络基础概念、数据库基础查询。
- 工具：远程协助、截图标注、文档编写、PPT汇报。

## 期望岗位

优先：软件实施、技术支持、IT运维/信息管理员、陶瓷电商运营助理、数据助理、项目助理。

暂不优先：高强度纯销售、长期夜班、重体力仓储、要求多年经验的高级开发岗。
"""


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.border = border
            cell.alignment = wrap
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    ws.freeze_panes = "A2"


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "投递前检查"
    ws.append(["模块", "必须完成"])
    for row in CHECKLIST_ROWS:
        ws.append(row)
    style_sheet(ws)
    set_widths(ws, {"A": 18, "B": 80})

    ws = wb.create_sheet("平台账号准备")
    ws.append(["平台", "入口", "投递特点", "必须准备", "容易卡住", "建议动作"])
    for item in PLATFORM_ROWS:
        ws.append([item[k] for k in ["平台", "入口", "投递特点", "必须准备", "容易卡住", "建议动作"]])
    style_sheet(ws)
    set_widths(ws, {"A": 16, "B": 34, "C": 42, "D": 48, "E": 48, "F": 56})

    ws = wb.create_sheet("投递跟踪")
    ws.append(TRACKER_HEADERS)
    for _ in range(30):
        ws.append([""] * len(TRACKER_HEADERS))
    style_sheet(ws)
    set_widths(ws, {
        "A": 14, "B": 14, "C": 24, "D": 26, "E": 12, "F": 18, "G": 42,
        "H": 14, "I": 16, "J": 12, "K": 16, "L": 16, "M": 12, "N": 12,
        "O": 18, "P": 24, "Q": 14, "R": 36,
    })

    ws = wb.create_sheet("话术模板")
    ws.append(["场景", "话术"])
    for item in SCRIPT_ROWS:
        ws.append([item["场景"], item["话术"]])
    style_sheet(ws)
    set_widths(ws, {"A": 24, "B": 100})

    ws = wb.create_sheet("一周投递节奏")
    ws.append(["日期", "任务", "数量目标", "完成情况"])
    rows = [
        ("第1天", "完善4个平台在线简历，准备PDF/DOCX附件简历", "账号4个，简历2版", ""),
        ("第2天", "投陶瓷电商运营、数据助理、文创商品助理", "8-12个", ""),
        ("第3天", "投技术支持、IT运维、信息管理员", "8-12个", ""),
        ("第4天", "投软件实施、项目助理、信息化交付", "8-12个", ""),
        ("第5天", "复盘已读/回复，优化简历关键词和开场白", "跟进10个", ""),
        ("第6天", "补投企业官网/公众号/本地招聘群", "5-8个", ""),
        ("第7天", "整理面试问题，筛掉高风险岗位", "复盘1次", ""),
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    set_widths(ws, {"A": 12, "B": 70, "C": 24, "D": 28})

    if os.path.exists(LOCAL_CHANNELS_PATH):
        with open(LOCAL_CHANNELS_PATH, "r", encoding="utf-8") as f:
            channels = json.load(f)
        ws = wb.create_sheet("本地小众渠道")
        ws.append(["渠道", "类型", "入口", "适合原因", "怎么用", "适合岗位", "注意事项"])
        for item in channels:
            ws.append([
                item.get("channel", ""),
                item.get("type", ""),
                item.get("url", ""),
                item.get("fit", ""),
                item.get("how_to_use", ""),
                item.get("best_for", ""),
                item.get("watch_out", ""),
            ])
        style_sheet(ws)
        set_widths(ws, {"A": 28, "B": 18, "C": 46, "D": 50, "E": 58, "F": 48, "G": 54})

    wb.save(path)


def main():
    configure_console_encoding()
    xlsx_path = os.path.join(BASE_DIR, "简历投递准备与跟踪表.xlsx")
    md_path = os.path.join(BASE_DIR, "简历投递版模板.md")
    create_workbook(xlsx_path)
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(RESUME_MARKDOWN)

    print("已生成投递准备包：")
    print(xlsx_path)
    print(md_path)
    print("生成时间：" + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


if __name__ == "__main__":
    main()
