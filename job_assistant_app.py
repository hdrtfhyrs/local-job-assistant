import datetime as dt
import glob
import json
import os
import subprocess
import sys
import tempfile
import threading
import time
import tkinter as tk
import webbrowser
from tkinter import messagebox, simpledialog, ttk


BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ---- 作者信息（本工具原创作者，请勿删除署名）----
AUTHOR_NAME = "1285452862"
BILIBILI_URL = "https://space.bilibili.com/352671558"
GITHUB_URL = "https://github.com/hdrtfhyrs/local-job-assistant"


def resolve_python_exe():
    # 优先用项目自带 venv312（保证子进程脚本有 requests/pandas/openpyxl 等依赖，
    # 无论 GUI 本身是用哪个 Python 启动的）
    venv_py = os.path.join(BASE_DIR, "venv312", "Scripts", "python.exe")
    if os.path.exists(venv_py):
        return venv_py
    exe = sys.executable
    if exe.lower().endswith("pythonw.exe"):
        candidate = os.path.join(os.path.dirname(exe), "python.exe")
        if os.path.exists(candidate):
            return candidate
    return exe


PYTHON_EXE = resolve_python_exe()

PLATFORMS = {
    "boss": {"name": "BOSS直聘", "home_url": "https://www.zhipin.com/"},
    "51job": {"name": "前程无忧", "home_url": "https://www.51job.com/"},
    "zhaopin": {"name": "智联招聘", "home_url": "https://www.zhaopin.com/"},
    "liepin": {"name": "猎聘", "home_url": "https://www.liepin.com/zhaogongzuo/"},
    "lagou": {"name": "拉勾招聘", "home_url": "https://www.lagou.com/wn/"},
}

STATUS_ITEMS = [
    ("research", "方向研究", "先看适合投什么"),
    ("package", "投递准备", "简历和跟踪表"),
    ("jobs", "岗位采集", "采集结果 Excel"),
    ("handoff", "精排候选", "规则筛选产出"),
    ("analysis", "AI精排结果", "本地qwen自动"),
    ("final", "最终推荐", "可直接打开"),
]

DEFAULT_KEYWORDS = "软件实施工程师,技术支持工程师,IT运维工程师,陶瓷电商运营助理,数据助理"

APP_BG = "#f6efe9"
SIDEBAR_BG = "#f3ebe5"
SIDEBAR_ACTIVE = "#e8ded7"
SIDEBAR_HOVER = "#eee5df"
CONTENT_BG = "#fffdfb"
CARD_BG = "#ffffff"
BORDER = "#e2ddd7"
TEXT = "#2c2621"
MUTED = "#7a716a"
PRIMARY = "#2563eb"
PRIMARY_DARK = "#1d4ed8"
ACCENT = "#0f766e"
WARNING = "#c2410c"
OK = "#15803d"


class JobAssistantApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("入门求职工作台")
        _sw, _sh = self.winfo_screenwidth(), self.winfo_screenheight()
        _w, _h = 1180, min(760, _sh - 80)
        self.geometry(f"{_w}x{_h}+{max(0, (_sw - _w) // 2)}+20")
        self.minsize(1120, 720)
        self.configure(bg=APP_BG)

        self.running = False
        self.task_buttons = []
        self.status_value_labels = {}
        self.keywords_text = None

        self.city_var = tk.StringVar(value="景德镇")
        self.platform_var = tk.StringVar(value="51job")
        self.pages_var = tk.IntVar(value=1)
        self.scrolls_var = tk.IntVar(value=4)
        self.top_var = tk.IntVar(value=10)
        self.match_top_var = tk.IntVar(value=25)
        self.wait_var = tk.IntVar(value=45)
        self.manual_var = tk.BooleanVar(value=False)
        self.snippets_var = tk.BooleanVar(value=False)
        self.fast_var = tk.BooleanVar(value=False)
        self.strict_boss_var = tk.BooleanVar(value=False)
        self.all_output_var = tk.BooleanVar(value=True)
        self.match_backend_var = tk.BooleanVar(value=False)
        self.status_vars = {key: tk.StringVar(value="未检查") for key, _, _ in STATUS_ITEMS}
        self.current_task_var = tk.StringVar(value="准备就绪")
        self.execution_state_var = tk.StringVar(value="空闲")
        self.execution_detail_var = tk.StringVar(value="等待操作：请选择左侧流程按钮。")
        self.next_step_var = tk.StringVar(value="第一次用？点左边「①  AI 帮我选方向」开始，然后照 ②③④⑤ 往下点。")
        self.current_command_title = None
        self.current_command_context = None
        self.current_on_success = None
        self.execution_progress = None
        self.execution_progress_visible = False
        self.execution_badge = None

        self._setup_styles()
        self._build_ui()
        self.refresh_status()

    def _setup_styles(self):
        self.option_add("*Font", ("Microsoft YaHei UI", 10))
        self.option_add("*TCombobox*Listbox.font", ("Microsoft YaHei UI", 10))

        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure(".", background=CONTENT_BG, foreground=TEXT, font=("Microsoft YaHei UI", 10))
        style.configure("TEntry", padding=(8, 5), fieldbackground="#ffffff", bordercolor=BORDER, lightcolor=BORDER, darkcolor=BORDER)
        style.configure("TCombobox", padding=(8, 5), fieldbackground="#ffffff", bordercolor=BORDER, lightcolor=BORDER, darkcolor=BORDER)
        style.configure("TSpinbox", padding=(6, 4), fieldbackground="#ffffff", bordercolor=BORDER, lightcolor=BORDER, darkcolor=BORDER)
        style.configure("TCheckbutton", background=CARD_BG, foreground=TEXT)
        style.map("TCheckbutton", background=[("active", CARD_BG)])
        style.configure("TButton", padding=(10, 7), font=("Microsoft YaHei UI", 10))
        style.configure(
            "Clean.Vertical.TScrollbar",
            gripcount=0,
            width=8,
            troughcolor=CONTENT_BG,
            background="#d8d2cc",
            bordercolor=CONTENT_BG,
            lightcolor="#d8d2cc",
            darkcolor="#d8d2cc",
            arrowcolor="#d8d2cc",
            relief=tk.FLAT,
        )
        style.map(
            "Clean.Vertical.TScrollbar",
            background=[("active", "#c8c1bb"), ("pressed", "#b8b0aa")],
            arrowcolor=[("active", "#c8c1bb"), ("pressed", "#b8b0aa")],
        )
        style.configure(
            "Execution.Horizontal.TProgressbar",
            troughcolor="#f4f0eb",
            background=PRIMARY,
            bordercolor=BORDER,
            lightcolor=PRIMARY,
            darkcolor=PRIMARY,
        )

    def _build_ui(self):
        shell = tk.Frame(self, bg=APP_BG)
        shell.pack(fill=tk.BOTH, expand=True)

        self.sidebar = tk.Frame(shell, bg=SIDEBAR_BG, width=300)
        self.sidebar.pack(side=tk.LEFT, fill=tk.Y)
        self.sidebar.pack_propagate(False)

        self.content = tk.Frame(shell, bg=CONTENT_BG)
        self.content.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self._build_sidebar()
        self._build_workspace()

    def _build_sidebar(self):
        canvas = tk.Canvas(self.sidebar, bg=SIDEBAR_BG, highlightthickness=0, bd=0)
        vbar = ttk.Scrollbar(self.sidebar, orient=tk.VERTICAL, command=canvas.yview,
                             style="Clean.Vertical.TScrollbar")
        canvas.configure(yscrollcommand=vbar.set)
        vbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        body = tk.Frame(canvas, bg=SIDEBAR_BG)
        body_id = canvas.create_window((0, 0), window=body, anchor="nw")
        self.sidebar_body = body
        body.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(body_id, width=e.width))
        canvas.bind("<Enter>", lambda e: canvas.bind_all(
            "<MouseWheel>", lambda ev: canvas.yview_scroll(int(-1 * (ev.delta / 120)), "units")))

        top = tk.Frame(body, bg=SIDEBAR_BG)
        top.pack(fill=tk.X, padx=18, pady=(18, 12))
        tk.Label(top, text="求职工作台", bg=SIDEBAR_BG, fg=TEXT, font=("Microsoft YaHei UI", 15, "bold")).pack(anchor=tk.W)
        tk.Label(top, text="本地流程助手", bg=SIDEBAR_BG, fg=MUTED, font=("Microsoft YaHei UI", 9)).pack(anchor=tk.W, pady=(3, 0))

        # ---- 常用（日常就这几个）----
        self._sidebar_section("常用")
        self._sidebar_button("🧑  我的资料（先填这个）", self.setup_profile, managed=True, active=True)
        self._sidebar_button("🌐  打开网站 · 登录 BOSS", self.launch_chrome, active=True)
        self._sidebar_button("🚀  一键全自动", self.run_all, managed=True, active=True, accent=True)
        self._sidebar_button("📊  看推荐结果（网页）", self.build_webview, managed=True, active=True)
        self._sidebar_button("🔔  看谁回信了", self.check_replies, managed=True)

        # ---- 高级（默认折叠，点开才显示）----
        self._adv_shown = False
        self.adv_toggle = tk.Button(
            body, text="⚙  分步 / 高级 / 文件      ▶", command=self._toggle_advanced,
            anchor="w", relief=tk.FLAT, bd=0, padx=16, pady=9,
            bg=SIDEBAR_BG, fg=MUTED, activebackground=SIDEBAR_HOVER, activeforeground=TEXT,
            font=("Microsoft YaHei UI", 10), cursor="hand2",
        )
        self.adv_toggle.pack(fill=tk.X, padx=10, pady=(16, 1))
        self.advanced_frame = tk.Frame(body, bg=SIDEBAR_BG)
        self._build_advanced(self.advanced_frame)  # 先建好内容，但不显示

        footer = tk.Frame(body, bg=SIDEBAR_BG)
        footer.pack(fill=tk.X, padx=18, pady=16)
        tk.Label(footer, text="本工具作者（原创）", bg=SIDEBAR_BG, fg=MUTED,
                 font=("Microsoft YaHei UI", 9)).pack(anchor=tk.W)
        tk.Label(footer, text="B站 @" + AUTHOR_NAME, bg=SIDEBAR_BG, fg=TEXT,
                 font=("Microsoft YaHei UI", 10, "bold")).pack(anchor=tk.W, pady=(3, 0))
        self._footer_link(footer, "▶ 我的 B 站主页", BILIBILI_URL)
        self._footer_link(footer, "★ GitHub 开源地址", GITHUB_URL)
        self._adv_footer = footer
        self._adv_canvas = canvas
        canvas.after(150, lambda: canvas.yview_moveto(0))

    def _footer_link(self, parent, text, url):
        lbl = tk.Label(parent, text=text, bg=SIDEBAR_BG, fg=ACCENT, cursor="hand2",
                       font=("Microsoft YaHei UI", 9, "underline"))
        lbl.pack(anchor=tk.W, pady=(4, 0))
        lbl.bind("<Button-1>", lambda e: webbrowser.open(url))

    def _build_advanced(self, p):
        self._sidebar_section("分步流程（手动）", parent=p)
        self._sidebar_button("①  AI 帮我选方向", self.career_advice, managed=True, parent=p)
        self._sidebar_button("③  开始抓岗位", self.scrape_jobs, managed=True, parent=p)
        self._sidebar_button("④  AI 筛选并出推荐表", self.one_click_filter, managed=True, parent=p)
        self._sidebar_button("🔍  投递预演（不真发）", self.auto_apply_dry, managed=True, parent=p)
        self._sidebar_button("🚀  仅投递（不爬取）", self.auto_apply, managed=True, parent=p)

        self._sidebar_section("其它工具", parent=p)
        self._sidebar_button("📍 本地特殊渠道", self.show_local_channels, parent=p)
        self._sidebar_button("职业方向研究表", self.generate_research, managed=True, parent=p)
        self._sidebar_button("简历投递准备包", self.generate_application_package, managed=True, parent=p)
        self._sidebar_button("单步·规则筛选", self.generate_handoff, managed=True, parent=p)
        self._sidebar_button("单步·AI 精排", self.local_ai_match, managed=True, parent=p)
        self._sidebar_button("单步·生成推荐表", self.merge_analysis, managed=True, parent=p)

        self._sidebar_section("文件", parent=p)
        self._sidebar_button("最新岗位表", self.open_latest_jobs, parent=p)
        self._sidebar_button("最终推荐表", self.open_latest_final, parent=p)
        self._sidebar_button("精排文件夹", self.open_inbox, parent=p)
        self._sidebar_button("项目文件夹", self.open_folder, parent=p)

    def _toggle_advanced(self):
        if self._adv_shown:
            self.advanced_frame.pack_forget()
            self.adv_toggle.config(text="⚙  分步 / 高级 / 文件      ▶")
        else:
            self.advanced_frame.pack(fill=tk.X, before=self._adv_footer)
            self.adv_toggle.config(text="⚙  分步 / 高级 / 文件      ▼")
        self._adv_shown = not self._adv_shown
        self.advanced_frame.update_idletasks()
        try:
            self._adv_canvas.configure(scrollregion=self._adv_canvas.bbox("all"))
        except Exception:
            pass

    def _sidebar_section(self, title, parent=None):
        tk.Label(
            parent or self.sidebar_body,
            text=title,
            bg=SIDEBAR_BG,
            fg="#9a8f87",
            font=("Microsoft YaHei UI", 9, "bold"),
        ).pack(anchor=tk.W, padx=18, pady=(14, 6))

    def _sidebar_button(self, text, command, managed=False, active=False, accent=False, parent=None):
        bg = SIDEBAR_ACTIVE if active else SIDEBAR_BG
        fg = ACCENT if accent else TEXT
        button = tk.Button(
            parent or self.sidebar_body,
            text=text,
            command=command,
            anchor="w",
            relief=tk.FLAT,
            bd=0,
            padx=16,
            pady=9,
            bg=bg,
            fg=fg,
            activebackground=SIDEBAR_HOVER,
            activeforeground=fg,
            font=("Microsoft YaHei UI", 10, "bold" if active or accent else "normal"),
            cursor="hand2",
        )
        button.pack(fill=tk.X, padx=10, pady=1)
        if managed:
            self.task_buttons.append(button)
        return button

    def _build_workspace(self):
        header = tk.Frame(self.content, bg=CONTENT_BG)
        header.pack(fill=tk.X, padx=38, pady=(22, 8))

        title_row = tk.Frame(header, bg=CONTENT_BG)
        title_row.pack(fill=tk.X)
        tk.Label(
            title_row,
            text="景德镇求职助手",
            bg=CONTENT_BG,
            fg=TEXT,
            font=("Microsoft YaHei UI", 16, "bold"),
        ).pack(side=tk.LEFT)
        self._pill(title_row, "当前任务  " + self.current_task_var.get(), side=tk.RIGHT)

        tk.Label(
            header,
            text="第一次用：按左边 ① → ⑤ 顺序点就行；每点一步，就看下面的「下一步提示」。",
            bg=CONTENT_BG,
            fg=MUTED,
            font=("Microsoft YaHei UI", 10),
        ).pack(anchor=tk.W, pady=(8, 0))

        scroll_shell = tk.Frame(self.content, bg=CONTENT_BG)
        scroll_shell.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(scroll_shell, bg=CONTENT_BG, highlightthickness=0, bd=0)
        scroll_indicator = tk.Canvas(scroll_shell, width=10, bg=CONTENT_BG, highlightthickness=0, bd=0)
        scroll_thumb = scroll_indicator.create_rectangle(3, 0, 7, 40, fill="#d3ccc6", outline="")

        def set_scroll_indicator(first, last):
            first = float(first)
            last = float(last)
            height = max(scroll_indicator.winfo_height(), 1)
            top = max(4, int(first * height))
            bottom = min(height - 4, int(last * height))
            if bottom - top < 34:
                bottom = min(height - 4, top + 34)
            scroll_indicator.coords(scroll_thumb, 3, top, 7, bottom)
            if first <= 0 and last >= 1:
                scroll_indicator.itemconfigure(scroll_thumb, state=tk.HIDDEN)
            else:
                scroll_indicator.itemconfigure(scroll_thumb, state=tk.NORMAL)

        canvas.configure(yscrollcommand=set_scroll_indicator)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_indicator.pack(side=tk.RIGHT, fill=tk.Y)

        scroll_body = tk.Frame(canvas, bg=CONTENT_BG)
        body_window = canvas.create_window((0, 0), window=scroll_body, anchor="nw")
        body = tk.Frame(scroll_body, bg=CONTENT_BG)
        body.pack(fill=tk.BOTH, expand=True, padx=38, pady=(0, 18))

        def sync_scroll_region(_event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def sync_body_width(event):
            canvas.itemconfigure(body_window, width=event.width)

        def on_mousewheel(event):
            steps = max(-3, min(3, int(-1 * (event.delta / 120))))
            canvas.yview_scroll(steps, "units")

        def indicator_moveto(event):
            height = max(scroll_indicator.winfo_height(), 1)
            canvas.yview_moveto(min(max(event.y / height, 0), 1))

        scroll_body.bind("<Configure>", sync_scroll_region)
        canvas.bind("<Configure>", sync_body_width)
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", on_mousewheel))
        scroll_indicator.bind("<Button-1>", indicator_moveto)
        scroll_indicator.bind("<B1-Motion>", indicator_moveto)

        self._build_execution_card(body)
        self._build_status_strip(body)
        self._build_settings_card(body)
        self._build_log_card(body)
        canvas.after(200, lambda: canvas.yview_moveto(0))

    def _pill(self, parent, text, side=tk.LEFT):
        label = tk.Label(
            parent,
            text=text,
            bg="#f4f0eb",
            fg="#5f564f",
            padx=12,
            pady=6,
            font=("Microsoft YaHei UI", 9, "bold"),
        )
        label.pack(side=side)
        self.task_pill = label

    def _card(self, parent, padx=16, pady=14):
        outer = tk.Frame(parent, bg=BORDER)
        inner = tk.Frame(outer, bg=CARD_BG)
        inner.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
        inner.configure(padx=padx, pady=pady)
        return outer, inner

    def _build_execution_card(self, parent):
        outer, card = self._card(parent, padx=12, pady=10)
        outer.pack(fill=tk.X, pady=(0, 8))

        header = tk.Frame(card, bg=CARD_BG)
        header.pack(fill=tk.X, pady=(0, 6))
        tk.Label(header, text="执行状态", bg=CARD_BG, fg=TEXT, font=("Microsoft YaHei UI", 12, "bold")).pack(side=tk.LEFT)
        self.execution_badge = tk.Label(
            header,
            textvariable=self.execution_state_var,
            bg="#f4f0eb",
            fg="#5f564f",
            padx=12,
            pady=5,
            font=("Microsoft YaHei UI", 9, "bold"),
        )
        self.execution_badge.pack(side=tk.RIGHT)

        body = tk.Frame(card, bg=CARD_BG)
        body.pack(fill=tk.X)
        body.columnconfigure(0, weight=2)
        body.columnconfigure(1, weight=3)

        current_box = tk.Frame(body, bg="#faf8f6", highlightbackground="#e7e1dc", highlightthickness=1)
        current_box.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        tk.Label(current_box, text="现在正在做", bg="#faf8f6", fg=MUTED, font=("Microsoft YaHei UI", 9)).pack(anchor=tk.W, padx=12, pady=(7, 1))
        tk.Label(
            current_box,
            textvariable=self.execution_detail_var,
            bg="#faf8f6",
            fg=TEXT,
            wraplength=390,
            justify=tk.LEFT,
            font=("Microsoft YaHei UI", 10, "bold"),
        ).pack(anchor=tk.W, fill=tk.X, padx=12, pady=(0, 8))

        next_box = tk.Frame(body, bg="#f7fbff", highlightbackground="#d9e7fb", highlightthickness=1)
        next_box.grid(row=0, column=1, sticky="nsew")
        tk.Label(next_box, text="下一步提示", bg="#f7fbff", fg=PRIMARY_DARK, font=("Microsoft YaHei UI", 9, "bold")).pack(anchor=tk.W, padx=12, pady=(7, 1))
        tk.Label(
            next_box,
            textvariable=self.next_step_var,
            bg="#f7fbff",
            fg=TEXT,
            wraplength=520,
            justify=tk.LEFT,
            font=("Microsoft YaHei UI", 10),
        ).pack(anchor=tk.W, fill=tk.X, padx=12, pady=(0, 8))

        self.execution_progress = ttk.Progressbar(
            card,
            mode="indeterminate",
            style="Execution.Horizontal.TProgressbar",
        )

    def _build_status_strip(self, parent):
        outer, card = self._card(parent, padx=12, pady=10)
        outer.pack(fill=tk.X, pady=(0, 8))

        header = tk.Frame(card, bg=CARD_BG)
        header.pack(fill=tk.X, pady=(0, 8))
        tk.Label(header, text="流程状态", bg=CARD_BG, fg=TEXT, font=("Microsoft YaHei UI", 12, "bold")).pack(side=tk.LEFT)
        tk.Button(
            header,
            text="刷新状态",
            command=self.refresh_status,
            relief=tk.FLAT,
            bg="#f5f2ef",
            fg=TEXT,
            padx=12,
            pady=6,
            activebackground="#ece7e2",
            cursor="hand2",
        ).pack(side=tk.RIGHT)

        grid = tk.Frame(card, bg=CARD_BG)
        grid.pack(fill=tk.X)
        for index, (key, label, hint) in enumerate(STATUS_ITEMS):
            col = index
            item = tk.Frame(grid, bg="#faf8f6", highlightbackground="#e7e1dc", highlightthickness=1)
            item.grid(row=0, column=col, sticky="ew", padx=(0 if col == 0 else 8, 0))
            grid.columnconfigure(col, weight=1)
            tk.Label(item, text=label, bg="#faf8f6", fg=TEXT, font=("Microsoft YaHei UI", 9, "bold")).pack(anchor=tk.W, padx=10, pady=(7, 1))
            value = tk.Label(item, textvariable=self.status_vars[key], bg="#faf8f6", fg=WARNING, font=("Microsoft YaHei UI", 9, "bold"))
            value.pack(anchor=tk.W, padx=10)
            tk.Label(item, text=hint, bg="#faf8f6", fg=MUTED, font=("Microsoft YaHei UI", 8)).pack(anchor=tk.W, padx=10, pady=(1, 7))
            self.status_value_labels[key] = value

    def _build_settings_card(self, parent):
        outer, card = self._card(parent, padx=16, pady=12)
        outer.pack(fill=tk.X, pady=(0, 8))

        tk.Label(card, text="配置与关键词", bg=CARD_BG, fg=TEXT, font=("Microsoft YaHei UI", 12, "bold")).pack(anchor=tk.W)
        tk.Label(card, text="关键词框可以一行一个，也可以用逗号分隔。", bg=CARD_BG, fg=MUTED).pack(anchor=tk.W, pady=(2, 8))

        content = tk.Frame(card, bg=CARD_BG)
        content.pack(fill=tk.BOTH, expand=True)
        content.columnconfigure(0, minsize=380)
        content.columnconfigure(1, weight=1)
        content.rowconfigure(0, weight=1)

        left = tk.Frame(content, bg=CARD_BG)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 22))
        left.columnconfigure(1, weight=1)
        left.columnconfigure(3, weight=1)
        right = tk.Frame(content, bg=CARD_BG)
        right.grid(row=0, column=1, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)

        tk.Label(left, text="城市", bg=CARD_BG, fg=TEXT).grid(row=0, column=0, sticky="w", pady=(0, 8), padx=(0, 8))
        ttk.Entry(left, textvariable=self.city_var).grid(row=0, column=1, columnspan=3, sticky="ew", pady=(0, 6))

        tk.Label(left, text="平台", bg=CARD_BG, fg=TEXT).grid(row=1, column=0, sticky="w", pady=6, padx=(0, 8))
        ttk.Combobox(
            left,
            textvariable=self.platform_var,
            values=list(PLATFORMS.keys()),
            state="readonly",
        ).grid(row=1, column=1, columnspan=3, sticky="ew", pady=6)

        spin_fields = [
            ("采集页数", self.pages_var, 1, 10),
            ("滚动次数", self.scrolls_var, 1, 12),
            ("研究方向", self.top_var, 3, 20),
            ("精排岗位", self.match_top_var, 5, 80),
            ("等待秒数", self.wait_var, 0, 300),
        ]
        for index, (label, var, start, end) in enumerate(spin_fields):
            row, col = divmod(index, 2)
            grid_row = row + 2
            label_col = col * 2
            tk.Label(left, text=label, bg=CARD_BG, fg=TEXT).grid(row=grid_row, column=label_col, sticky="w", pady=6, padx=(0, 8))
            ttk.Spinbox(left, from_=start, to=end, textvariable=var, width=9).grid(row=grid_row, column=label_col + 1, sticky="w", pady=6, padx=(0, 18))

        options = tk.Frame(left, bg=CARD_BG)
        options.grid(row=5, column=0, columnspan=4, sticky="ew", pady=(8, 0))
        ttk.Checkbutton(options, text="人工搜索/过验证后采集", variable=self.manual_var).pack(anchor=tk.W, pady=1)
        ttk.Checkbutton(options, text="研究表收集搜索摘要", variable=self.snippets_var).pack(anchor=tk.W, pady=1)
        ttk.Checkbutton(options, text="快速模式", variable=self.fast_var).pack(anchor=tk.W, pady=1)
        ttk.Checkbutton(options, text="BOSS 严格过滤", variable=self.strict_boss_var).pack(anchor=tk.W, pady=1)
        ttk.Checkbutton(options, text="导出全部规则打分表", variable=self.all_output_var).pack(anchor=tk.W, pady=1)
        ttk.Checkbutton(options, text="精排用 Claude（最准·耗额度；不勾=本地 qwen 免费）", variable=self.match_backend_var).pack(anchor=tk.W, pady=1)

        tk.Label(right, text="岗位关键词", bg=CARD_BG, fg=TEXT, font=("Microsoft YaHei UI", 10, "bold")).grid(row=0, column=0, sticky="w")
        text_wrap = tk.Frame(right, bg="#ede7e1", highlightthickness=0)
        text_wrap.grid(row=1, column=0, sticky="nsew", pady=(7, 8))
        text_wrap.columnconfigure(0, weight=1)
        text_wrap.rowconfigure(0, weight=1)
        self.keywords_text = tk.Text(
            text_wrap,
            height=6,
            wrap=tk.WORD,
            bg="#fbfaf8",
            fg=TEXT,
            insertbackground=PRIMARY,
            relief=tk.FLAT,
            padx=12,
            pady=10,
            font=("Microsoft YaHei UI", 10),
        )
        self.keywords_text.grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        self.keywords_text.insert("1.0", DEFAULT_KEYWORDS.replace(",", "\n"))

        keyword_bar = tk.Frame(right, bg=CARD_BG)
        keyword_bar.grid(row=2, column=0, sticky="ew")
        tk.Label(keyword_bar, text="建议先用 3-6 个核心方向跑通流程。", bg=CARD_BG, fg=MUTED, font=("Microsoft YaHei UI", 9)).pack(side=tk.LEFT)
        self._small_button(keyword_bar, "恢复默认", self.reset_keywords).pack(side=tk.RIGHT)

    def _form_row(self, parent, label, widget):
        row = tk.Frame(parent, bg=CARD_BG)
        row.pack(fill=tk.X, pady=5)
        tk.Label(row, text=label, bg=CARD_BG, fg=TEXT, width=8, anchor="w").pack(side=tk.LEFT)
        widget.pack(side=tk.LEFT, fill=tk.X, expand=True)

    def _build_log_card(self, parent):
        outer, card = self._card(parent, padx=16, pady=12)
        outer.pack(fill=tk.X)

        header = tk.Frame(card, bg=CARD_BG)
        header.pack(fill=tk.X, pady=(0, 8))
        tk.Label(header, text="运行日志", bg=CARD_BG, fg=TEXT, font=("Microsoft YaHei UI", 12, "bold")).pack(side=tk.LEFT)
        self._small_button(header, "清空", self.clear_log).pack(side=tk.RIGHT)

        text_wrap = tk.Frame(card, bg="#fbfaf8", highlightbackground="#ede7e1", highlightthickness=1)
        text_wrap.pack(fill=tk.BOTH, expand=True)
        self.log_text = tk.Text(
            text_wrap,
            wrap=tk.WORD,
            height=8,
            bg="#fbfaf8",
            fg=TEXT,
            insertbackground=PRIMARY,
            relief=tk.FLAT,
            padx=14,
            pady=12,
            spacing1=3,
            spacing3=3,
            font=("Microsoft YaHei UI", 10),
        )
        scrollbar = ttk.Scrollbar(text_wrap, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.log("欢迎使用求职助手。第一次用：从左边「①  AI 帮我选方向」开始，照 ①→⑤ 顺序点就行。")
        self.log("①选方向(AI推荐+自动填词) → ②打开网站登录 → ③抓岗位 → ④AI筛选出表 → ⑤看结果")
        self.log("提示：②之后请在弹出的 Chrome 里登录 BOSS、过验证码（软件不会绕过验证）。")

    def _small_button(self, parent, text, command):
        return tk.Button(
            parent,
            text=text,
            command=command,
            relief=tk.FLAT,
            bd=0,
            bg="#f5f2ef",
            fg=TEXT,
            activebackground="#ece7e2",
            padx=12,
            pady=6,
            cursor="hand2",
            font=("Microsoft YaHei UI", 9),
        )

    def _int_value(self, var, default, minimum=None):
        try:
            value = int(var.get())
        except (tk.TclError, ValueError):
            value = default
        if minimum is not None:
            value = max(minimum, value)
        return value

    def _platform_key(self):
        key = self.platform_var.get().strip()
        return key if key in PLATFORMS else "51job"

    def get_keywords(self):
        if not self.keywords_text:
            return DEFAULT_KEYWORDS
        raw = self.keywords_text.get("1.0", tk.END).strip()
        parts = []
        for item in raw.replace("，", ",").replace("、", ",").replace("\n", ",").split(","):
            keyword = item.strip()
            if keyword and keyword not in parts:
                parts.append(keyword)
        return ",".join(parts)

    def reset_keywords(self):
        if not self.keywords_text:
            return
        self.keywords_text.delete("1.0", tk.END)
        self.keywords_text.insert("1.0", DEFAULT_KEYWORDS.replace(",", "\n"))

    def _latest_file(self, *patterns):
        files = []
        for pattern in patterns:
            files.extend(glob.glob(os.path.join(BASE_DIR, pattern)))
        files = [p for p in files if os.path.isfile(p) and not os.path.basename(p).startswith("~$")]
        if not files:
            return None
        return max(files, key=os.path.getmtime)

    def _brief_file_status(self, path):
        if not path:
            return "未生成"
        stamp = os.path.getmtime(path)
        text = dt.datetime.fromtimestamp(stamp).strftime("%m-%d %H:%M")
        return f"已有 {text}"

    def refresh_status(self):
        research = self._latest_file("*职业方向推荐*.xlsx", "*口碑线索*.xlsx")
        package = os.path.join(BASE_DIR, "简历投递准备与跟踪表.xlsx")
        jobs = self._latest_file("*岗位采集*.xlsx", "*IT过渡岗位推荐*.xlsx", "*自定义岗位搜集*.xlsx", "*本科软件工程入门岗位*.xlsx")
        handoff = os.path.join(BASE_DIR, "analysis_inbox", "AI分析任务.md")
        analysis = os.path.join(BASE_DIR, "analysis_inbox", "analysis_result.json")
        final = self._latest_file("*岗位精排推荐*.xlsx")
        if not final:
            final = self._latest_file("*岗位规则推荐*.xlsx")

        self.status_vars["research"].set(self._brief_file_status(research))
        self.status_vars["package"].set(self._brief_file_status(package if os.path.exists(package) else None))
        self.status_vars["jobs"].set(self._brief_file_status(jobs))
        self.status_vars["handoff"].set(self._brief_file_status(handoff if os.path.exists(handoff) else None))
        self.status_vars["analysis"].set(self._brief_file_status(analysis if os.path.exists(analysis) else None))
        self.status_vars["final"].set(self._brief_file_status(final))
        self._refresh_status_styles()

    def _refresh_status_styles(self):
        for key, label in self.status_value_labels.items():
            label.configure(fg=OK if self.status_vars[key].get().startswith("已有") else WARNING)

    def _set_current_task(self, text):
        self.current_task_var.set(text)
        if hasattr(self, "task_pill"):
            self.task_pill.configure(text="当前任务  " + text)

    def _set_execution_state(self, state, detail=None, next_step=None, running=False):
        self.execution_state_var.set(state)
        if detail is not None:
            self.execution_detail_var.set(detail)
        if next_step is not None:
            self.next_step_var.set(next_step)

        colors = {
            "空闲": ("#f4f0eb", "#5f564f"),
            "正在运行": ("#dbeafe", PRIMARY_DARK),
            "已完成": ("#dcfce7", OK),
            "采集偏少": ("#ffedd5", WARNING),
            "未完成": ("#ffedd5", WARNING),
            "失败": ("#fee2e2", "#b91c1c"),
            "需人工处理": ("#ffedd5", WARNING),
        }
        bg, fg = colors.get(state, ("#f4f0eb", "#5f564f"))
        if self.execution_badge is not None:
            self.execution_badge.configure(bg=bg, fg=fg)

        if self.execution_progress is not None:
            if running:
                if not self.execution_progress_visible:
                    self.execution_progress.pack(fill=tk.X, pady=(10, 0))
                    self.execution_progress_visible = True
                self.execution_progress.start(12)
            else:
                self.execution_progress.stop()
                if self.execution_progress_visible:
                    self.execution_progress.pack_forget()
                    self.execution_progress_visible = False

    def _next_step_after(self, title, code):
        if code != 0:
            return "这一步没有成功。请先看下面运行日志最后几行，再重新执行当前按钮。"
        if title == "智能推荐方向":
            return "已生成推荐方向、并把关键词自动填进采集框。下一步：点“启动调试 Chrome”登录，再点“采集岗位”。"
        if title == "生成职业方向研究表":
            return "下一步：点“生成投递准备包”，或者先点“启动调试 Chrome”登录招聘平台。"
        if title == "生成简历投递准备包":
            return "下一步：点“启动调试 Chrome”，登录或完成验证后再点“采集岗位”。"
        if title.startswith("采集岗位"):
            return "下一步：点左侧“规则筛选”，把采集到的岗位整理成可精排列表。"
        if title == "规则筛选并生成AI精排任务":
            return "下一步：点“本地AI精排”，用本机 qwen 自动精排并直接出表（无需AI/VPN）。"
        if title == "生成最终推荐表":
            return "下一步：点左侧“最终推荐表”打开结果。"
        return "下一步：按左侧流程继续执行。"

    def _file_signature(self, path):
        if not path or not os.path.exists(path):
            return None
        stat = os.stat(path)
        return (stat.st_mtime_ns, stat.st_size)

    def _paths_for_patterns(self, patterns):
        paths = []
        for pattern in patterns:
            if any(ch in pattern for ch in "*?[]"):
                paths.extend(glob.glob(os.path.join(BASE_DIR, pattern)))
            else:
                paths.append(os.path.join(BASE_DIR, pattern))
        clean = []
        for path in paths:
            if os.path.isfile(path) and not os.path.basename(path).startswith("~$") and path not in clean:
                clean.append(path)
        return clean

    def _snapshot_for_patterns(self, patterns):
        return {path: self._file_signature(path) for path in self._paths_for_patterns(patterns)}

    def _changed_paths(self, patterns, before):
        changed = []
        for path in self._paths_for_patterns(patterns):
            signature = self._file_signature(path)
            if signature and before.get(path) != signature:
                changed.append(path)
        changed.sort(key=os.path.getmtime, reverse=True)
        return changed

    def _command_patterns(self, title):
        if title == "生成职业方向研究表":
            return ["*职业方向推荐*.xlsx", "*口碑线索*.xlsx"]
        if title == "生成简历投递准备包":
            return ["简历投递准备与跟踪表.xlsx", "简历投递版模板.md"]
        if title.startswith("采集岗位"):
            return ["*岗位采集*.xlsx", "*IT过渡岗位推荐*.xlsx", "*自定义岗位搜集*.xlsx", "*本科软件工程入门岗位*.xlsx"]
        if title == "规则筛选并生成AI精排任务":
            return [os.path.join("analysis_inbox", "pending_jobs.json"), os.path.join("analysis_inbox", "AI分析任务.md"), "规则打分_全部.xlsx"]
        if title == "生成最终推荐表":
            return ["*岗位精排推荐*.xlsx", "*岗位规则推荐*.xlsx"]
        return []

    def _build_command_context(self, title):
        patterns = self._command_patterns(title)
        return {
            "title": title,
            "started_at": time.time(),
            "patterns": patterns,
            "before": self._snapshot_for_patterns(patterns),
        }

    def _pending_job_count(self):
        path = os.path.join(BASE_DIR, "analysis_inbox", "pending_jobs.json")
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            return 0
        jobs = data.get("岗位", []) if isinstance(data, dict) else []
        return len(jobs) if isinstance(jobs, list) else 0

    def _analysis_result_count(self):
        path = os.path.join(BASE_DIR, "analysis_inbox", "analysis_result.json")
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            return 0
        if isinstance(data, dict):
            data = data.get("results") or data.get("岗位") or []
        return len(data) if isinstance(data, list) else 0

    def _xlsx_stats(self, path):
        stats = {"rows": 0, "detail_rows": 0}
        if not path or not os.path.exists(path):
            return stats
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            stats["rows"] = max(ws.max_row - 1, 0)
            detail_col = None
            for idx, header in enumerate(headers, start=1):
                if header == "岗位详情":
                    detail_col = idx
                    break
            if detail_col:
                count = 0
                for row in ws.iter_rows(min_row=2, min_col=detail_col, max_col=detail_col, values_only=True):
                    value = str(row[0] or "").strip()
                    if value:
                        count += 1
                stats["detail_rows"] = count
        except Exception as exc:
            self.log(f"读取 Excel 统计失败：{exc}")
        return stats

    def _expected_job_floor(self):
        keywords = max(len([item for item in self.get_keywords().split(",") if item.strip()]), 1)
        pages = self._int_value(self.pages_var, 1, 1)
        if self._platform_key() == "boss":
            return min(max(keywords * pages * 8, 20), 120)
        return min(max(keywords * pages * 6, 15), 80)

    def _validate_command_completion(self, title, context):
        before = (context or {}).get("before", {})

        if title == "生成职业方向研究表":
            changed = self._changed_paths(["*职业方向推荐*.xlsx", "*口碑线索*.xlsx"], before)
            if changed:
                return {"ok": True, "state": "已完成", "detail": f"已生成研究表：{os.path.basename(changed[0])}"}
            return {"ok": False, "detail": "脚本结束了，但没有检测到本次新生成的研究表。", "next_step": "请查看日志最后几行；如果被登录、网络或浏览器卡住，需要处理后重试。"}

        if title == "生成简历投递准备包":
            required = ["简历投递准备与跟踪表.xlsx", "简历投递版模板.md"]
            changed = self._changed_paths(required, before)
            missing = [name for name in required if not os.path.exists(os.path.join(BASE_DIR, name))]
            if not missing and len(changed) == len(required):
                return {"ok": True, "state": "已完成", "detail": "已更新简历投递准备包和投递模板。"}
            return {"ok": False, "detail": "脚本结束了，但投递准备包没有完整更新。", "next_step": "请查看日志最后几行，再重新点击“生成投递准备包”。"}

        if title.startswith("采集岗位"):
            changed = self._changed_paths(["*岗位采集*.xlsx", "*IT过渡岗位推荐*.xlsx", "*自定义岗位搜集*.xlsx", "*本科软件工程入门岗位*.xlsx"], before)
            if changed:
                stats = self._xlsx_stats(changed[0])
                expected = self._expected_job_floor()
                detail = f"采到 {stats['rows']} 条，详情 {stats['detail_rows']} 条：{os.path.basename(changed[0])}"
                if stats["rows"] < expected:
                    return {
                        "ok": True,
                        "state": "采集偏少",
                        "detail": detail,
                        "next_step": "这次结果明显偏少。请确认 BOSS 在搜索结果列表页、城市正确、没有验证；也可以关闭“快速模式”后重试。",
                    }
                if self._platform_key() == "boss" and stats["detail_rows"] < stats["rows"]:
                    return {
                        "ok": True,
                        "state": "采集偏少",
                        "detail": detail,
                        "next_step": "岗位表已生成，但详情没有全抓完。建议重试一次，期间不要操作 Chrome。",
                    }
                return {"ok": True, "state": "已完成", "detail": detail}
            return {"ok": False, "detail": "采集脚本结束了，但没有检测到新的岗位 Excel。", "next_step": "大概率是页面未登录、验证没过、城市/关键词没有结果。请在 Chrome 里处理后重试采集。"}

        if title == "规则筛选并生成AI精排任务":
            required = [os.path.join("analysis_inbox", "pending_jobs.json"), os.path.join("analysis_inbox", "AI分析任务.md")]
            changed = self._changed_paths(required, before)
            job_count = self._pending_job_count()
            if len(changed) == len(required) and job_count > 0:
                return {"ok": True, "state": "已完成", "detail": f"已生成AI任务：{job_count} 个候选岗位。"}
            if job_count <= 0:
                return {"ok": False, "detail": "规则筛选结束了，但没有生成可交给AI的候选岗位。", "next_step": "请先确认岗位采集表里有数据，再重新运行“规则筛选”。"}
            return {"ok": False, "detail": "规则筛选脚本结束了，但AI任务文件不是本次新生成的。", "next_step": "请查看日志最后几行；必要时重新运行“规则筛选”。"}

        if title == "生成最终推荐表":
            changed_precise = self._changed_paths(["*岗位精排推荐*.xlsx"], before)
            changed_rule = self._changed_paths(["*岗位规则推荐*.xlsx"], before)
            result_count = self._analysis_result_count()
            if changed_precise and result_count > 0:
                return {"ok": True, "state": "已完成", "detail": f"已生成AI精排推荐表：{os.path.basename(changed_precise[0])}"}
            if changed_rule:
                return {
                    "ok": True,
                    "state": "需人工处理",
                    "detail": f"已生成规则兜底推荐：{os.path.basename(changed_rule[0])}，还不是AI精排结果。",
                    "next_step": "下一步：点“本地AI精排”，自动精排后会再生成精排表。",
                }
            return {"ok": False, "detail": "脚本结束了，但没有检测到本次新生成的最终推荐表。", "next_step": "请先运行“规则筛选”；如果已有AI结果，请确认 analysis_result.json 放在 analysis_inbox 里。"}

        return {"ok": True, "state": "已完成", "detail": f"已完成：{title}"}

    def log(self, message):
        text = str(message).rstrip()
        self.log_text.insert(tk.END, text + "\n")
        self.log_text.see(tk.END)
        self._update_execution_from_log(text)
        self.update_idletasks()

    def _update_execution_from_log(self, text):
        if not self.running or not self.current_command_title:
            return
        if self.current_command_title.startswith("采集岗位"):
            if "打开搜索页：" in text:
                self.execution_detail_var.set(text.strip())
            elif "详情采集目标：" in text:
                self.execution_detail_var.set(text.strip())
                self.next_step_var.set("正在逐个打开岗位详情页抓取内容，Chrome 停在详情页是正常现象。")
            elif "正在抓取:" in text and "/" in text:
                self.execution_detail_var.set(text.strip())
                self.next_step_var.set("正在抓详情页内容，请不要操作 Chrome。")
            elif "全部完成！共保存" in text:
                self.execution_detail_var.set(text.strip())

    def clear_log(self):
        self.log_text.delete("1.0", tk.END)

    def set_running(self, running):
        self.running = running
        state = tk.DISABLED if running else tk.NORMAL
        for button in self.task_buttons:
            button.configure(state=state)

    def run_command(self, args, title, on_success=None):
        if self.running:
            messagebox.showinfo("正在运行", "已有任务正在运行，请等它结束。")
            return

        self.current_command_title = title
        self.current_command_context = self._build_command_context(title)
        self.current_on_success = on_success
        self.set_running(True)
        self._set_current_task(title)
        self._set_execution_state(
            "正在运行",
            f"正在执行：{title}",
            "请等待日志继续输出。任务运行期间，左侧流程按钮会暂时禁用。",
            running=True,
        )
        self.log("")
        self.log("=" * 72)
        self.log(title)
        self.log("命令：" + subprocess.list2cmdline(args))

        thread = threading.Thread(target=self._run_command_worker, args=(args, on_success), daemon=True)
        thread.start()

    def _run_command_worker(self, args, on_success):
        code = None
        try:
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            process = subprocess.Popen(
                args,
                cwd=BASE_DIR,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                env=env,
            )
            assert process.stdout is not None
            for line in process.stdout:
                self.after(0, self.log, line.rstrip())
            code = process.wait()
            self.after(0, self.log, f"任务结束，退出码：{code}")
        except Exception as exc:
            self.after(0, self.log, f"任务失败：{exc}")
        finally:
            self.after(0, self._finish_command, code)

    def _finish_command(self, code):
        self.set_running(False)
        self.refresh_status()
        title = self.current_command_title or "上一个任务"
        if code == 0:
            validation = self._validate_command_completion(title, self.current_command_context)
            if validation.get("ok"):
                self._set_current_task("准备就绪")
                if self.current_on_success:
                    self.current_on_success()
                self._set_execution_state(
                    validation.get("state", "已完成"),
                    validation.get("detail", f"已完成：{title}"),
                    validation.get("next_step", self._next_step_after(title, code)),
                    running=False,
                )
            else:
                self._set_current_task("上一步未完成")
                self._set_execution_state(
                    validation.get("state", "未完成"),
                    validation.get("detail", f"{title} 没有生成预期结果。"),
                    validation.get("next_step", "请查看日志最后几行，再重新执行当前步骤。"),
                    running=False,
                )
        else:
            self._set_current_task("上一个任务需要检查日志")
            code_text = "未知错误" if code is None else f"退出码：{code}"
            self._set_execution_state("失败", f"{title} 没有完成（{code_text}）", self._next_step_after(title, code), running=False)
        self.current_command_title = None
        self.current_command_context = None
        self.current_on_success = None

    def find_chrome(self):
        local_app_data = os.environ.get("LocalAppData", "")
        candidates = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.join(local_app_data, r"Google\Chrome\Application\chrome.exe"),
        ]
        for path in candidates:
            if path and os.path.exists(path):
                return path
        return None

    def find_antigravity(self):
        local_app_data = os.environ.get("LocalAppData", "")
        candidates = [
            os.path.join(local_app_data, r"Programs\antigravity\Antigravity.exe"),
            os.path.join(local_app_data, r"Programs\Antigravity IDE\Antigravity IDE.exe"),
            r"C:\Program Files\Antigravity\Antigravity.exe",
            r"C:\Program Files\Antigravity IDE\Antigravity IDE.exe",
        ]
        for path in candidates:
            if path and os.path.exists(path):
                return path
        return None

    def launch_chrome(self):
        chrome = self.find_chrome()
        if not chrome:
            messagebox.showerror("未找到 Chrome", "没有找到 chrome.exe。请安装 Chrome，或手动用调试端口 9222 启动浏览器。")
            return

        platform = PLATFORMS[self._platform_key()]
        profile_dir = os.path.join(tempfile.gettempdir(), "boss_scraper_debug")
        args = [
            chrome,
            "--remote-debugging-port=9222",
            f"--user-data-dir={profile_dir}",
            "--no-first-run",
            "--no-default-browser-check",
            # 禁用 Chrome 后台/遮蔽节流:否则窗口一最小化或被挡，抓取页就停渲染、
            # 滚动加载不触发，导致抓不到。CalculateNativeWinOcclusion 是最关键的那一个。
            "--disable-backgrounding-occluded-windows",
            "--disable-renderer-backgrounding",
            "--disable-background-timer-throttling",
            "--disable-features=CalculateNativeWinOcclusion",
            platform["home_url"],
        ]
        try:
            subprocess.Popen(args, cwd=BASE_DIR)
        except Exception as exc:
            messagebox.showerror("启动失败", str(exc))
            self._set_execution_state("失败", "Chrome 没有启动成功。", "请确认电脑已安装 Chrome，或手动用 9222 调试端口启动。")
            return
        self._set_current_task("等待登录或验证")
        self._set_execution_state(
            "需人工处理",
            f"已打开调试 Chrome：{platform['name']}",
            "请在新打开的 Chrome 里登录、切换城市或完成验证；处理好后回到这里点“采集岗位”。",
        )
        self.log(f"已打开调试 Chrome：{platform['name']}。请先登录或完成验证，再回到工作台采集。")

    def career_advice(self):
        if self.running:
            messagebox.showinfo("正在运行", "已有任务在跑，请等它结束。")
            return
        info = self._ask_profile_dialog()
        if info is None:
            self.log("已取消「AI 帮我选方向」。")
            return
        self._save_profile(info)
        self.log("AI 正在联网核实各方向的真实评价（待遇/加班/坑），约 3-6 分钟，请耐心等、别关窗口…")

        def after_success():
            self._apply_recommended_keywords()
        self.run_command([PYTHON_EXE, "career_advisor.py", "--top", "6"], "智能推荐方向", on_success=after_success)

    def _load_profile_dict(self):
        try:
            with open(os.path.join(BASE_DIR, "match_profile.json"), "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}

    def _split_keywords(self, s):
        for sep in ("，", "、", "；", ";", ","):
            s = s.replace(sep, ",")
        return [x.strip() for x in s.split(",") if x.strip()]

    def _ask_profile_dialog(self):
        prof = self._load_profile_dict()
        dlg = tk.Toplevel(self)
        dlg.title("先说说你的情况（AI 据此帮你选方向）")
        dlg.configure(bg=CARD_BG)
        _dw, _dh = 580, min(620, self.winfo_screenheight() - 80)
        _dx = max(0, (self.winfo_screenwidth() - _dw) // 2)
        _dy = max(10, (self.winfo_screenheight() - _dh) // 2)
        dlg.geometry(f"{_dw}x{_dh}+{_dx}+{_dy}")
        dlg.minsize(480, 420)
        dlg.transient(self)
        dlg.grab_set()
        dlg.lift()
        dlg.focus_force()
        dlg.attributes("-topmost", True)
        dlg.after(400, lambda: dlg.attributes("-topmost", False))

        tk.Label(dlg, text="把你的真实情况填一下，AI 会据此推荐最适合你的求职方向。\n（已预填默认值，改一改、或直接用都行）",
                 bg=CARD_BG, fg=MUTED, font=("Microsoft YaHei UI", 10),
                 wraplength=540, justify=tk.LEFT).pack(anchor=tk.W, padx=18, pady=(14, 4))

        state = {"ok": False, "data": None}

        def on_ok():
            # 必须在 dlg.destroy() 之前读取所有控件的值
            # （尤其 tk.Text，控件销毁后再 .get() 会报 invalid command name）
            try:
                salary = float(salary_var.get())
            except Exception:
                salary = prof.get("salary_floor_k", 3.5)
            state["data"] = {
                "city": city_var.get().strip() or "景德镇",
                "background": bg_var.get().strip(),
                "have": self._split_keywords(skills_var.get()),
                "want_to_grow": self._split_keywords(grow_var.get()),
                "salary_floor_k": salary,
                "accept_relocation": relo_var.get(),
                "extra_notes": notes_text.get("1.0", tk.END).strip(),
            }
            state["ok"] = True
            dlg.destroy()

        # 按钮固定在底部，永远可见（先 pack）
        btns = tk.Frame(dlg, bg=CARD_BG)
        btns.pack(side=tk.BOTTOM, fill=tk.X, padx=18, pady=14)
        tk.Button(btns, text="开始让 AI 推荐", command=on_ok, bg=PRIMARY, fg="#ffffff",
                  relief=tk.FLAT, padx=16, pady=8, cursor="hand2",
                  font=("Microsoft YaHei UI", 10, "bold")).pack(side=tk.RIGHT)
        tk.Button(btns, text="取消", command=dlg.destroy, bg="#eee5df", fg=TEXT,
                  relief=tk.FLAT, padx=16, pady=8, cursor="hand2").pack(side=tk.RIGHT, padx=(0, 10))

        # 字段区可滚动
        fcanvas = tk.Canvas(dlg, bg=CARD_BG, highlightthickness=0, bd=0)
        fbar = ttk.Scrollbar(dlg, orient=tk.VERTICAL, command=fcanvas.yview,
                             style="Clean.Vertical.TScrollbar")
        fcanvas.configure(yscrollcommand=fbar.set)
        fbar.pack(side=tk.RIGHT, fill=tk.Y)
        fcanvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        form = tk.Frame(fcanvas, bg=CARD_BG)
        form_id = fcanvas.create_window((0, 0), window=form, anchor="nw")
        form.bind("<Configure>", lambda e: fcanvas.configure(scrollregion=fcanvas.bbox("all")))
        fcanvas.bind("<Configure>", lambda e: fcanvas.itemconfigure(form_id, width=e.width))
        fcanvas.bind("<Enter>", lambda e: fcanvas.bind_all(
            "<MouseWheel>", lambda ev: fcanvas.yview_scroll(int(-1 * (ev.delta / 120)), "units")))

        def field(label, hint=""):
            tk.Label(form, text=label, bg=CARD_BG, fg=TEXT,
                     font=("Microsoft YaHei UI", 10, "bold")).pack(anchor=tk.W, padx=18, pady=(10, 2))
            if hint:
                tk.Label(form, text=hint, bg=CARD_BG, fg=MUTED,
                         font=("Microsoft YaHei UI", 9)).pack(anchor=tk.W, padx=18)

        field("你在哪个城市找工作？")
        city_var = tk.StringVar(value=prof.get("city", "景德镇"))
        ttk.Entry(form, textvariable=city_var).pack(fill=tk.X, padx=18)

        field("你的学历、专业、技术水平？", "例：普通本科 软件工程，技术基础一般")
        bg_var = tk.StringVar(value=prof.get("background", "普通本科 软件工程，技术基础一般"))
        ttk.Entry(form, textvariable=bg_var).pack(fill=tk.X, padx=18)

        field("你已经会哪些技能？", "逗号分隔，例：SQL, Excel, Linux, Python基础")
        skills_var = tk.StringVar(value="，".join(prof.get("skills", {}).get("have", [])))
        ttk.Entry(form, textvariable=skills_var).pack(fill=tk.X, padx=18)

        field("有没有想做 / 感兴趣的方向？", "可留空，让 AI 全权推荐；例：运维, 电商运营, 数据")
        grow_var = tk.StringVar(value="，".join(prof.get("skills", {}).get("want_to_grow", [])))
        ttk.Entry(form, textvariable=grow_var).pack(fill=tk.X, padx=18)

        field("期望月薪下限（单位 K，即千元）？", "例：3.5 表示 3500 元")
        salary_var = tk.StringVar(value=str(prof.get("salary_floor_k", 3.5)))
        ttk.Entry(form, textvariable=salary_var).pack(fill=tk.X, padx=18)

        relo_var = tk.BooleanVar(value=bool(prof.get("accept_relocation", False)))
        ttk.Checkbutton(form, text="能接受长期出差 / 驻场（不勾 = 只要本地、不驻场）",
                        variable=relo_var).pack(anchor=tk.W, padx=18, pady=(12, 0))

        field("其他要求或顾虑？", "例：工时别太长、不要体力活、想稳定、最好双休…")
        notes_text = tk.Text(form, height=3, wrap=tk.WORD, relief=tk.FLAT, bg="#fbfaf8",
                             font=("Microsoft YaHei UI", 10), padx=8, pady=6)
        notes_text.pack(fill=tk.X, padx=18, pady=(2, 12))
        notes_text.insert("1.0", prof.get("extra_notes", ""))

        self.wait_window(dlg)
        try:
            self.unbind_all("<MouseWheel>")
        except Exception:
            pass
        return state["data"] if state["ok"] else None

    def _save_profile(self, info):
        path = os.path.join(BASE_DIR, "match_profile.json")
        prof = self._load_profile_dict()
        prof["city"] = info["city"]
        prof["background"] = info["background"]
        prof["salary_floor_k"] = info["salary_floor_k"]
        prof["accept_relocation"] = info["accept_relocation"]
        prof["extra_notes"] = info["extra_notes"]
        prof.setdefault("skills", {})
        prof["skills"]["have"] = info["have"]
        prof["skills"]["want_to_grow"] = info["want_to_grow"]
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(prof, f, ensure_ascii=False, indent=2)
            self.log(f"已记下你的情况：{info['city']}｜{info['background']}｜底薪{info['salary_floor_k']}K｜"
                     f"{'接受' if info['accept_relocation'] else '不接受'}驻场")
        except Exception as exc:
            self.log(f"保存情况失败：{exc}")

    def _apply_recommended_keywords(self):
        path = os.path.join(BASE_DIR, "recommended_directions.json")
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as exc:
            self.log(f"读取推荐方向失败：{exc}")
            return
        kws = data.get("all_keywords", [])
        if kws and self.keywords_text:
            self.keywords_text.delete("1.0", tk.END)
            self.keywords_text.insert("1.0", "\n".join(kws))
            self.log(f"已把 {len(kws)} 个推荐关键词自动填入采集框：{('、'.join(kws))}")
            self.log("下一步：点“启动调试 Chrome”登录，再点“采集岗位”。")
        self._show_directions_window(data)

    def _show_directions_window(self, data):
        recs = data.get("recommendations", [])
        if not recs:
            return
        win = tk.Toplevel(self)
        win.title("AI 推荐的求职方向（看完点关闭）")
        win.configure(bg=CARD_BG)
        _w = 640
        _h = min(620, self.winfo_screenheight() - 100)
        _x = max(0, (self.winfo_screenwidth() - _w) // 2)
        _y = max(10, (self.winfo_screenheight() - _h) // 2)
        win.geometry(f"{_w}x{_h}+{_x}+{_y}")
        win.transient(self)
        win.lift()
        win.focus_force()
        win.attributes("-topmost", True)
        win.after(400, lambda: win.attributes("-topmost", False))

        bar = tk.Frame(win, bg=CARD_BG)
        bar.pack(side=tk.BOTTOM, fill=tk.X, padx=16, pady=12)
        tk.Button(bar, text="知道了，关闭", command=win.destroy, bg=PRIMARY, fg="#ffffff",
                  relief=tk.FLAT, padx=16, pady=8, cursor="hand2",
                  font=("Microsoft YaHei UI", 10, "bold")).pack(side=tk.RIGHT)

        wrap = tk.Frame(win, bg=CARD_BG)
        wrap.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=2, pady=2)
        txt = tk.Text(wrap, wrap=tk.WORD, relief=tk.FLAT, bg="#fbfaf8", fg=TEXT,
                      font=("Microsoft YaHei UI", 10), padx=16, pady=12)
        sb = ttk.Scrollbar(wrap, orient=tk.VERTICAL, command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        order = {"高": 0, "中": 1, "低": 2}
        rlist = sorted(recs, key=lambda r: order.get(str(r.get("priority", "中")), 1))
        txt.insert(tk.END, f"AI（{data.get('model', '')}）按你的情况推荐了 {len(rlist)} 个方向：\n\n")
        for i, r in enumerate(rlist, 1):
            txt.insert(tk.END, f"{i}. {r.get('direction', '')}　【优先级：{r.get('priority', '')}】\n")
            if r.get("why_fit"):
                txt.insert(tk.END, f"    · 为什么适合你：{r['why_fit']}\n")
            if r.get("local_opportunity"):
                txt.insert(tk.END, f"    · 本地机会：{r['local_opportunity']}\n")
            if r.get("real_review"):
                txt.insert(tk.END, f"    · 🔍 网上真实评价：{r['real_review']}\n")
            if r.get("search_keywords"):
                txt.insert(tk.END, f"    · 去招聘网站搜：{'、'.join(r['search_keywords'])}\n")
            if r.get("learn_first"):
                txt.insert(tk.END, f"    · 上手前先补：{'、'.join(r['learn_first'])}\n")
            txt.insert(tk.END, "\n")
        txt.insert(tk.END, "这些方向的关键词已自动填进左边采集框；下一步点「② 打开网站并登录」。")
        txt.configure(state=tk.DISABLED)

    def show_local_channels(self):
        """展示景德镇本地特殊招聘渠道(大平台之外的机会)。app 内窗口、带滚动。"""
        import json as _json
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "local_niche_channels.json")
        try:
            with open(path, "r", encoding="utf-8") as f:
                channels = _json.load(f)
        except Exception as e:
            self.log(f"打开本地渠道失败：{e}")
            return
        if not channels:
            self.log("暂无本地渠道数据。")
            return

        win = tk.Toplevel(self)
        win.title("景德镇本地特殊招聘渠道（大平台之外的机会）")
        win.configure(bg=CARD_BG)
        _w = 680
        _h = min(640, self.winfo_screenheight() - 100)
        _x = max(0, (self.winfo_screenwidth() - _w) // 2)
        _y = max(10, (self.winfo_screenheight() - _h) // 2)
        win.geometry(f"{_w}x{_h}+{_x}+{_y}")
        win.transient(self)
        win.lift()
        win.focus_force()
        win.attributes("-topmost", True)
        win.after(400, lambda: win.attributes("-topmost", False))

        bar = tk.Frame(win, bg=CARD_BG)
        bar.pack(side=tk.BOTTOM, fill=tk.X, padx=16, pady=12)
        tk.Button(bar, text="知道了，关闭", command=win.destroy, bg=PRIMARY, fg="#ffffff",
                  relief=tk.FLAT, padx=16, pady=8, cursor="hand2",
                  font=("Microsoft YaHei UI", 10, "bold")).pack(side=tk.RIGHT)

        wrap = tk.Frame(win, bg=CARD_BG)
        wrap.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=2, pady=2)
        txt = tk.Text(wrap, wrap=tk.WORD, relief=tk.FLAT, bg="#fbfaf8", fg=TEXT,
                      font=("Microsoft YaHei UI", 10), padx=16, pady=12)
        sb = ttk.Scrollbar(wrap, orient=tk.VERTICAL, command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        txt.insert(tk.END, "大平台(BOSS/智联)之外，景德镇本地这些渠道常有大平台搜不到的岗位，\n"
                           "尤其适合应届/入门、陶瓷电商、文旅、内部IT。建议每周固定翻一翻：\n\n")
        for i, c in enumerate(channels, 1):
            txt.insert(tk.END, f"{i}. {c.get('channel','')}　【{c.get('type','')}】\n")
            if c.get("url"):
                txt.insert(tk.END, f"    · 在哪：{c['url']}\n")
            if c.get("fit"):
                txt.insert(tk.END, f"    · 适合啥：{c['fit']}\n")
            if c.get("best_for"):
                txt.insert(tk.END, f"    · 重点找：{c['best_for']}\n")
            if c.get("how_to_use"):
                txt.insert(tk.END, f"    · 怎么用：{c['how_to_use']}\n")
            if c.get("watch_out"):
                txt.insert(tk.END, f"    · ⚠️ 注意：{c['watch_out']}\n")
            txt.insert(tk.END, "\n")
        txt.insert(tk.END, "提示：看到公司名先去『企查查/天眼查』查真伪，谨防收费培训、押金、模糊薪资、无公司主体。")
        txt.configure(state=tk.DISABLED)

    def generate_research(self):
        args = [PYTHON_EXE, "career_research_collector.py", "--top", str(self._int_value(self.top_var, 10, 1))]
        if self.snippets_var.get():
            args.append("--collect-snippets")
        self.run_command(args, "生成职业方向研究表")

    def generate_application_package(self):
        self.run_command([PYTHON_EXE, "application_package.py"], "生成简历投递准备包")

    def scrape_jobs(self):
        platform_key = self._platform_key()
        city = self.city_var.get().strip() or "景德镇"
        keywords = self.get_keywords()

        if platform_key == "boss":
            args = [
                PYTHON_EXE,
                "boss_scraper.py",
                "--city",
                city,
                "--max-pages",
                str(self._int_value(self.pages_var, 1, 1)),
                "--max-details",
                "0",
                "--skip-reputation",
                "--yes",
            ]
            if keywords:
                args.extend(["--keywords", keywords])
            if not self.strict_boss_var.get():
                args.append("--no-strict-match")
            if self.fast_var.get():
                args.append("--fast")
        else:
            args = [
                PYTHON_EXE,
                "scrape_beginner_jobs.py",
                "--platform",
                platform_key,
                "--city",
                city,
                "--pages",
                str(self._int_value(self.pages_var, 1, 1)),
                "--scrolls",
                str(self._int_value(self.scrolls_var, 4, 1)),
                "--wait-before-collect",
                str(self._int_value(self.wait_var, 0, 0)),
                "--yes",
            ]
            if keywords:
                args.extend(["--keywords", keywords])
            if self.manual_var.get():
                args.append("--manual")
            if self.fast_var.get():
                args.append("--fast")

        self.run_command(args, f"采集岗位：{PLATFORMS[platform_key]['name']}")

    def generate_handoff(self):
        args = [
            PYTHON_EXE,
            "job_matcher.py",
            "--top",
            str(self._int_value(self.match_top_var, 25, 1)),
        ]
        if self.all_output_var.get():
            args.append("--all-output")

        def after_success():
            self.log("已生成规则筛选结果。下一步点击“本地AI精排”，用本机 qwen 自动精排并出表（无需AI/VPN）。")

        self.run_command(args, "规则筛选并生成AI精排任务", on_success=after_success)

    def one_click_filter(self):
        latest = self._latest_file("*岗位采集*.xlsx", "*IT过渡岗位推荐*.xlsx",
                                   "*自定义岗位搜集*.xlsx", "*本科软件工程入门岗位*.xlsx")
        if not latest:
            messagebox.showinfo("还没有岗位数据", "请先点「③ 开始抓岗位」采集岗位，再来这一步。")
            self._set_execution_state("未完成", "还没有采集到的岗位表。",
                                      "下一步：先点「③ 开始抓岗位」采集。")
            return
        brain = "Claude(最准)" if self.match_backend_var.get() else "本地 qwen(免费)"
        self.log(f"开始一键筛选：规则筛选 → AI 精排（{brain}） → 出推荐表，请耐心等。")
        args = [PYTHON_EXE, "job_matcher.py", "--top",
                str(self._int_value(self.match_top_var, 25, 1))]
        if self.all_output_var.get():
            args.append("--all-output")

        def after_match():
            self.log("规则筛选完成，自动进入 AI 精排…")
            self.after(1200, self._auto_ai_after_match)

        self.run_command(args, "规则筛选并生成AI精排任务", on_success=after_match)

    def _auto_ai_after_match(self):
        if self.running:
            self.after(1200, self._auto_ai_after_match)
            return
        self.local_ai_match()

    def local_ai_match(self):
        pending = os.path.join(BASE_DIR, "analysis_inbox", "pending_jobs.json")
        if not os.path.exists(pending):
            messagebox.showinfo("还没有候选岗位", "请先点“规则筛选”，生成待精排的候选岗位。")
            self._set_current_task("缺少候选岗位")
            self._set_execution_state(
                "未完成",
                "还没有 pending_jobs.json。",
                "下一步：先点左侧“规则筛选”。",
            )
            return

        backend = "claude" if self.match_backend_var.get() else "ollama"
        brain = "Claude(最准)" if backend == "claude" else "本地qwen(免费)"
        self.log(f"精排大脑：{brain}")

        def after_success():
            self.log("AI 精排完成，正在自动生成最终推荐表……")
            self.after(1200, self._auto_merge_after_ai)

        self.run_command([PYTHON_EXE, "local_ai_matcher.py", "--backend", backend],
                         "本地AI精排", on_success=after_success)

    def _auto_merge_after_ai(self):
        if self.running:
            self.after(1200, self._auto_merge_after_ai)
            return
        self.merge_analysis()

    def merge_analysis(self):
        result_path = os.path.join(BASE_DIR, "analysis_inbox", "analysis_result.json")
        if not os.path.exists(result_path):
            self.log("未发现 analysis_result.json，将先生成规则分兜底推荐表；AI写回后可再次生成精排推荐表。")
        self.run_command([PYTHON_EXE, "merge_analysis.py"], "生成最终推荐表",
                         on_success=self._after_table_build_web)

    def _after_table_build_web(self):
        self.log("推荐表已生成，正在自动生成网页视图……")
        self.after(900, self._run_build_web)

    def _run_build_web(self):
        if self.running:
            self.after(900, self._run_build_web)
            return
        self.run_command([PYTHON_EXE, "build_web.py"], "生成推荐网页", on_success=self._open_webview)

    def open_path(self, path, missing_message):
        if path and os.path.exists(path):
            os.startfile(path)
            return True
        messagebox.showinfo("未找到文件", missing_message)
        return False

    def open_antigravity_task(self):
        path = os.path.join(BASE_DIR, "analysis_inbox", "AI分析任务.md")
        inbox = os.path.dirname(path)
        if not os.path.exists(path):
            messagebox.showinfo("还没有任务", "还没有生成AI精排任务。请先运行“规则筛选”。")
            self._set_current_task("缺少AI任务")
            self._set_execution_state(
                "未完成",
                "还没有AI任务文件。",
                "下一步：先点左侧“规则筛选”，生成 analysis_inbox/AI分析任务.md。",
            )
            return

        self.clipboard_clear()
        self.clipboard_append(path)

        antigravity = self.find_antigravity()
        launched_antigravity = False
        if antigravity:
            try:
                subprocess.Popen([antigravity, path], cwd=BASE_DIR)
                launched_antigravity = True
            except Exception as exc:
                self.log(f"启动 Antigravity 失败：{exc}")

        try:
            subprocess.Popen(["explorer", "/select,", path])
        except Exception as exc:
            self.log(f"打开精排文件夹失败：{exc}")

        self._set_current_task("等待AI写回")
        if launched_antigravity:
            detail = "已启动 Antigravity，并已把任务文件路径复制到剪贴板。"
        else:
            detail = "未自动找到或启动 Antigravity；已把任务文件路径复制到剪贴板，并打开精排文件夹。"
        self._set_execution_state(
            "需人工处理",
            detail,
            "请在 Antigravity 中打开/粘贴任务文件路径，按任务要求生成 analysis_inbox/analysis_result.json；写回后再点“生成最终推荐表”。",
        )
        self.log(f"AI任务文件：{path}")
        self.log("任务路径已复制到剪贴板。请让 Antigravity 按该 .md 文件写回 analysis_inbox/analysis_result.json。")

    def open_latest_jobs(self):
        path = self._latest_file("*岗位采集*.xlsx", "*IT过渡岗位推荐*.xlsx", "*自定义岗位搜集*.xlsx", "*本科软件工程入门岗位*.xlsx")
        self.open_path(path, "还没有岗位采集结果。请先运行“采集岗位”。")

    def open_latest_final(self):
        path = self._latest_file("*岗位精排推荐*.xlsx", "*岗位规则推荐*.xlsx")
        self.open_path(path, "还没有最终推荐表。请先运行“生成最终推荐表”。")

    def open_inbox(self):
        path = os.path.join(BASE_DIR, "analysis_inbox")
        os.makedirs(path, exist_ok=True)
        os.startfile(path)

    def open_folder(self):
        os.startfile(BASE_DIR)

    # ---------------- 个性化资料 ----------------
    def setup_profile(self):
        prof_path = os.path.join(BASE_DIR, "user_profile.json")
        cur = {}
        if os.path.exists(prof_path):
            try:
                with open(prof_path, "r", encoding="utf-8") as f:
                    cur = json.load(f)
            except Exception:
                cur = {}

        def ask(label, key, default):
            return simpledialog.askstring("个性化设置", label,
                                          initialvalue=str(cur.get(key, default)), parent=self)

        city = ask("你在哪个城市？", "city", "景德镇")
        if city is None:
            return
        edu = ask("最高学历？（如 本科/大专）", "education", "本科")
        bg = ask("一句话背景（专业/会什么技能）：", "background", "")
        tg = ask("想做哪些方向？空格隔开\n（如：电商运营 数据助理 行政）", "targets", "")
        av = ask("有什么不接受？（如：长期出差）", "avoid", "")
        sal = ask("薪资底线（K，如 3.5）：", "salary_floor_k", "3.5")
        try:
            float(sal)
        except (TypeError, ValueError):
            sal = "3.5"
        self.log("正在根据你的资料生成个性化投递规则（搜哪些岗位、跳过哪些不对口的）……")
        args = [PYTHON_EXE, "profile_engine.py",
                "--city", city or "景德镇", "--edu", edu or "本科",
                "--background", bg or "", "--targets", tg or "",
                "--avoid", av or "", "--salary", sal]
        self.run_command(args, "生成个性化画像")

    # ---------------- 一键全自动 ----------------
    def run_all(self):
        do_apply = True
        cfgp = os.path.join(BASE_DIR, "自动运行配置.json")
        if os.path.exists(cfgp):
            try:
                with open(cfgp, "r", encoding="utf-8") as f:
                    do_apply = json.load(f).get("do_apply", True)
            except Exception:
                pass
        tail = "→ 自动投递 → 查回信" if do_apply else "→ 查回信（投递已暂停，本次不投）"
        if not messagebox.askyesno(
            "一键全自动",
            "将自动跑完整套流程：\n爬岗位 → 筛选 → AI精排 → 出推荐表 → 出网页 " + tail + "。\n\n"
            "请先点【🌐 打开网站·登录BOSS】扫码登录。\n"
            "运行期间不要操作弹出的那个 Chrome 窗口。\n\n现在开始吗？",
        ):
            return
        self.log("开始一键全自动……投递时请勿操作那个调试 Chrome。")
        self.run_command([PYTHON_EXE, "auto_run_all.py"], "一键全自动")

    # ---------------- 自动投递 / 回信监控 / 网页 ----------------
    def auto_apply(self):
        cap = simpledialog.askinteger(
            "全自动投递",
            "今天最多投多少个？\n（建议 50~80；太多容易触发 BOSS 风控被封号）",
            initialvalue=60, minvalue=1, maxvalue=200, parent=self,
        )
        if cap is None:
            return
        if not messagebox.askyesno(
            "最后确认",
            f"将用你的 BOSS 账号【真的】给符合条件的岗位发招呼语，最多 {cap} 个。\n\n"
            "已自动带：每条随机间隔、去重(投过不再投)、避开'不建议/踩坑'岗位。\n\n"
            "请确保已点②打开网站并登录 BOSS。现在开始吗？",
        ):
            return
        self.log("开始全自动投递。期间请不要操作那个调试 Chrome 窗口。")
        self.run_command([PYTHON_EXE, "auto_apply.py", "--max", str(cap)], "全自动投递")

    def auto_apply_dry(self):
        self.log("投递预演：只定位'立即沟通'按钮、不会真的发，用来检查能不能正常投。")
        self.run_command([PYTHON_EXE, "auto_apply.py", "--dry-run"], "投递预演（不真发）")

    def check_replies(self):
        self.run_command([PYTHON_EXE, "reply_monitor.py"], "查看谁回信了",
                         on_success=self._open_reply_xlsx)

    def _open_reply_xlsx(self):
        path = os.path.join(BASE_DIR, "回信记录.xlsx")
        if os.path.exists(path):
            try:
                os.startfile(path)
            except Exception:
                pass

    def build_webview(self):
        self.run_command([PYTHON_EXE, "build_web.py"], "生成推荐网页",
                         on_success=self._open_webview)

    def _open_webview(self):
        path = os.path.join(os.path.expanduser("~"), "Desktop", "求职推荐_网页版.html")
        if os.path.exists(path):
            try:
                os.startfile(path)
            except Exception:
                pass


if __name__ == "__main__":
    app = JobAssistantApp()
    app.mainloop()
